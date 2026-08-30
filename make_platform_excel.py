# -*- coding: utf-8 -*-
"""
PS5 Completion Platform  ->  Excel mirror generator + live sync.

Builds a downloadable .xlsx that looks EXACTLY like the platform
(same tables, same colors, same layout) and keeps it in sync with the
online cloud state (platform_state.json on GitHub).

Any of the 3 accounts can, from any laptop:
  1. double-click SYNCHRONIZE.bat (desktop icon)
  2. type edits directly in the colored Excel
  3. double-click the icon again -> edits are pushed online, the latest
     state from the platform is pulled, and the file is updated.

Modes:
  python make_platform_excel.py --sync    push my edits + pull + reopen
  python make_platform_excel.py --pull    pull cloud only + rebuild + reopen
  python make_platform_excel.py --push    push my edits only (no rebuild)
  python make_platform_excel.py --build   rebuild from local index.html   [no fetch]
  python make_platform_excel.py --local   use LOCAL index.html if present (fallback then fetch)

Data source (any laptop, no local files needed):
  index.html            -> https://raw.githubusercontent.com/Mohamedgawad1/PS5-COMPLETION-PLATFORM/main/index.html
  platform_state.json   -> https://raw.githubusercontent.com/Mohamedgawad1/PS5-COMPLETION-PLATFORM/main/platform_state.json
"""
import base64
import json
import os
import re
import sys
import time
import urllib.request

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print('ERROR: openpyxl is not installed. Run:  pip install openpyxl')
    sys.exit(1)

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = HERE
XLSX = os.path.join(OUT_DIR, 'PS5 PLATFORM.xlsx')
SNAP = os.path.join(OUT_DIR, '_snapshot.json')
TOKEN = os.path.join(OUT_DIR, 'platform_token.txt')

REPO = 'Mohamedgawad1/PS5-COMPLETION-PLATFORM'
RAW_INDEX = f'https://raw.githubusercontent.com/{REPO}/main/index.html'
RAW_STATE = f'https://raw.githubusercontent.com/{REPO}/main/platform_state.json'
API_STATE = f'https://api.github.com/repos/{REPO}/contents/platform_state.json'

TITLE = 'PS5 COMPLETION PLATFORM'

# ------------------------------------------------------------------ colors --
# discipline light fills (RFC detail cells) + dark header fills
DC = {'B': '#e8d5c4', 'E': '#d4e6f1', 'H': '#d5f5e3', 'I': '#fdebd0',
      'M': '#fadbd8', 'P': '#e8daef', 'S': '#f9e79f', 'T': '#d6eaf8'}
DL = {'B': '#8b5e3c', 'E': '#1a5276', 'H': '#1e8449', 'I': '#b9770e',
      'M': '#922b21', 'P': '#6c3483', 'S': '#7d6608', 'T': '#1b4f72'}
RFK = ['B', 'E', 'H', 'I', 'M', 'P', 'S', 'T']

# cable discipline fills
CAB_DC = {'Electrical': '#d4e6f1', 'Instrument': '#fdebd0',
          'Telecom': '#d6eaf8', 'Trace Heating': '#fadbd8'}

# category fills
CAT_FILL = {'A': '#ffc7ce', 'B': '#ffeb9c', 'C': '#c6efce'}

# percent cell colors (matches pcol() in the platform)
PCOL_NUM = [(90, '#2e7d32'), (50, '#ed7d31'), (0, '#c00000'), (None, '#8a97a5')]

# RGB -> ARGB for openpyxl
def a(c):
    c = str(c or '').lstrip('#')
    return c if len(c) == 8 else 'FF' + c

def fill(c):
    return PatternFill(start_color=a(c), end_color=a(c), fill_type='solid')

NAVY = '#0b2239'
NAVY2 = '#10365e'
ACC = '#ed7d31'
GREEN = '#2e7d32'
ORANGE = '#ed7d31'
RED = '#c00000'
GRAY = '#8a97a5'
YELLOW = '#ffeb9c'
GREENC = '#c6efce'
REMARK = '#fff8e1'
BLACK = '#000000'
ZEBRA = '#f8fafd'
WHITE = '#FFFFFF'
LINK = '#0a58ca'

THIN = Side(style='thin', color='D0D7E2')
THICK_ACC = Side(style='thick', color='ED7D31')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FONT = 'Calibri'

# ----------------------------------------------------------------- helpers --
def fetch(url, timeout=90):
    req = urllib.request.Request(url, headers={'User-Agent': 'ps5-excel'})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read()


def split_array(line, name):
    """Find 'name=[...]' in a JS const line and return the JSON array."""
    idx = line.find(f'{name}=[')
    if idx < 0:
        return []
    i = idx + len(name) + 1
    depth = 0
    for j in range(i, len(line)):
        if line[j] == '[':
            depth += 1
        elif line[j] == ']':
            depth -= 1
            if depth == 0:
                return json.loads(line[i:j + 1])
    return []


def split_obj(line, name):
    idx = line.find(f'{name}=')
    if idx < 0:
        return {}
    i = idx + len(name) + 1
    depth = 0
    for j in range(i, len(line)):
        if line[j] == '{':
            depth += 1
        elif line[j] == '}':
            depth -= 1
            if depth == 0:
                return json.loads(line[i:j + 1])
    return {}


def parse_index(html, ovr_add=None, notes_add=None, pcol_add=None):
    """Return (SUBS,PUNCH,ITR,RFC,ITRT,PUNT,MILES,CABLES,OVR,NOTES,PCOL)."""
    lines = html.split('\n')
    data_line = next(ln for ln in lines if 'const SUBS=' in ln)
    SUBS = split_array(data_line, 'SUBS')
    PUNCH = split_array(data_line, 'PUNCH')
    ITR = split_array(data_line, 'ITR')
    RFC = split_array(data_line, 'RFC')
    ITRT = split_array(data_line, 'ITRT')
    PUNT = split_array(data_line, 'PUNT')
    MILES = split_array(data_line, 'MILES')
    CABLES = split_array(data_line, 'CABLES')

    ovr_line = next(ln for ln in lines if 'const NOTES=' in ln)
    OVR = split_obj(ovr_line, 'OVR')
    NOTES = split_obj(ovr_line, 'NOTES')
    PCOL = split_obj(ovr_line, 'PCOL')

    # cloud merge EXACTLY like the web (mergeCloud): cloud fills only where
    # local/OVR has nothing -> local wins.
    for sh, ids in (ovr_add or {}).items():
        for idv, cols in ids.items():
            OVR.setdefault(sh, {}).setdefault(idv, {})
            for c, v in cols.items():
                if OVR[sh][idv].get(c) in (None, '', None):
                    if OVR[sh][idv].get(c) is None:
                        OVR[sh][idv][c] = v
            # keep: only set when key missing
            for c, v in list(cols.items()):
                if c not in OVR[sh][idv]:
                    OVR[sh][idv][c] = v
    for sh, ids in (notes_add or {}).items():
        NOTES.setdefault(sh, {})
        for idv, v in ids.items():
            if idv not in NOTES[sh]:
                NOTES[sh][idv] = v
    for sh, ids in (pcol_add or {}).items():
        PCOL.setdefault(sh, {})
        for idv, v in ids.items():
            if idv not in PCOL[sh]:
                PCOL[sh][idv] = v
    return SUBS, PUNCH, ITR, RFC, ITRT, PUNT, MILES, CABLES, OVR, NOTES, PCOL


def v(ovr, sh, sid, col, default):
    o = (ovr.get(sh) or {}).get(sid)
    if o and col in o and o[col] not in (None, ''):
        return o[col]
    return default


def pct_color(p):
    for th, c in PCOL_NUM:
        if th is not None and p >= th:
            return c
    return GRAY


def norm(vv):
    if vv is None:
        return ''
    if isinstance(vv, float) and vv == int(vv):
        vv = int(vv)
    return str(vv).strip()


def done_state(s):
    s = str(s or '').strip()
    if re.search(r'to\s*be', s, re.I):
        return False
    return bool(re.match(r'^comp', s, re.I)) or s.lower() == 'closed'


# ------------------------------------------------------------- sheet writer --
class WB:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)

    def sheet(self, name):
        ws = self.wb.create_sheet(title=name[:31])
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 26
        return ws

    def title(self, ws, ncols, sub):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        cell = ws.cell(1, 1, value=f'{TITLE}  |  {sub}')
        cell.font = Font(name=FONT, size=13, bold=True, color=a(WHITE))
        cell.fill = fill(NAVY)
        cell.alignment = Alignment(vertical='center', horizontal='left',
                                   indent=1)
        for c in range(1, ncols + 1):
            ws.cell(1, c).fill = fill(NAVY)
            ws.cell(1, c).border = Border(left=THIN, right=THIN, top=THIN,
                                          bottom=THICK_ACC)

    def head(self, ws, row, values, spans=None, fills=None, height=30):
        """spans: {col_index(excel 1-based): colspan}; group header fill map."""
        c = 1
        for i, txt in enumerate(values, 1):
            if spans and i in spans:
                for _ in range(spans[i]):
                    cell = ws.cell(row, c + _)
                    if _ == 0:
                        cell.value = txt
                    cell.fill = fill((fills or {}).get(i, NAVY2))
                    cell.font = Font(name=FONT, size=8, bold=True,
                                      color=a(WHITE))
                    cell.alignment = Alignment(horizontal='center',
                                               vertical='center',
                                               wrap_text=True)
                    cell.border = BORDER
                ws.merge_cells(start_row=row, start_column=c,
                               end_row=row, end_column=c + spans[i] - 1)
                c += spans[i]
            else:
                cell = ws.cell(row, c, value=txt)
                cell.fill = fill((fills or {}).get(i, NAVY2))
                cell.font = Font(name=FONT, size=9, bold=True,
                                          color=a(WHITE))
                cell.alignment = Alignment(horizontal='center',
                                           vertical='center', wrap_text=True)
                cell.border = BORDER
                c += 1
        ws.row_dimensions[row].height = height
        ws.freeze_panes = ws.cell(row + 1, 1).coordinate
        return c - 1

    def widths(self, ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def finalize(self, ws, nrows, header_row, ncols):
        ws.auto_filter.ref = f'{get_column_letter(1)}{header_row}:' \
                             f'{get_column_letter(ncols)}{nrows}'


def done_style(ws, r, c, col=None, bold=False, italic=False, size=10,
               color=None, bg=None, wrap=False, center=False, b=True):
    cell = ws.cell(r, c)
    cell.font = Font(name=FONT, size=size, bold=bold, italic=italic,
                     color=a(color) if color else 'FF1A2A3A')
    if bg:
        cell.fill = fill(bg)
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical='top', indent=1)
    elif center:
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
    else:
        cell.alignment = Alignment(vertical='center', wrap_text=False)
    if b:
        cell.border = BORDER


PLATFORM = 'https://mohamedgawad1.github.io/PS5-COMPLETION-PLATFORM'


def link_cell(ws, r, c, text, url, bold=False, size=10, bg=None,
              fade=False, wrap=False, center=False):
    cell = ws.cell(r, c, value=text)
    cell.hyperlink = url
    cell.font = Font(name=FONT, size=size, bold=bold, italic=False,
                     color=a(fade and NAVY2 or LINK), underline='single')
    if bg:
        cell.fill = fill(bg)
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical='center', indent=1)
    elif center:
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
    else:
        cell.alignment = Alignment(vertical='center', wrap_text=False)
    cell.border = BORDER
    return cell


# ---------------------------------------------------------------- building --
def build_workbook(SUBS, PUNCH, ITR, RFC, ITRT, PUNT, MILES, CABLES,
                   OVR, NOTES, PCOL):
    w = WB()

    # ---------------- RFC PROGRESS ----------------
    ws = w.sheet('RFC PROGRESS')
    SH = 'RFC PROGRESS'
    # header col groups: 9 fixed + 8*(3) + 6 + 3 + 3 + 1 + 1 + 1 = 46
    fixed = ['SUB SYSTEM', 'PRIO', 'RFC BHM', 'RFC EIT', 'BASE', 'RECOV',
             'SIGNED', 'MILESTONE', 'TOT%']
    fixed_psh = {1: 'SUB', 2: 'PRIO', 3: 'RFC BHMPS', 4: 'RFC EIT',
                 5: 'Baseline', 6: 'Recovery', 7: 'SIGNED', 8: 'Milestone'}
    row1 = list(fixed)
    row1 += ['B', 'E', 'H', 'I', 'M', 'P', 'S', 'T']
    row1 += ['ITRs', 'CL', 'BAL', 'ITR%', 'CPP', 'EIT', 'EACP',
             'REMARK E', 'REMARK C', 'REMARK 1', 'WALKDOWN', '\u25cf', 'MY NOTES']
    # discipline group spans start at excel col 10
    grp_start = len(fixed) + 1
    spans = {grp_start + k: 3 for k in range(8)}
    fills = {grp_start + k: DL[RFK[k]] for k in range(8)}
    for k in range(8):          # color the two sub-heads T/CL/% too
        base = grp_start + k
        for off in range(1, 3):
            fills[base + off] = DL[RFK[k]]
    w.title(ws, 46, '2 \u00b7 RFC PROGRESS')
    w.head(ws, 2, row1, spans=spans, fills=fills, height=32)
    # second header row T/CL/%
    r2 = [''] * len(fixed)
    for k in range(8):
        r2 += ['T', 'CL', '%']
    r2 += [''] * 13
    cell_fills = {}
    for k in range(8):
        base = grp_start + k
        for off in range(3):
            cell_fills[base + off] = DL[RFK[k]]
    w.head(ws, 3, r2, fills=cell_fills, height=14)
    widths = [34, 6, 10, 10, 10, 10, 10, 16, 8]
    for k in range(8):
        widths += [6, 6, 6]
    widths += [8, 8, 8, 9, 7, 7, 7, 14, 14, 14, 14, 4, 22]

    r = 4
    note_fonts = []
    for x in RFC:
        sid = x['sid']
        name = x['name']
        sg = v(OVR, SH, sid, 'SIGNED', x['signed']).upper()
        tot = x['tot']
        present7 = (tot != '' and tot >= 100) or sg.startswith('COMPLETE') \
            or sg.startswith('PARTIAL')
        sbg = GREENC if sg.startswith('COMPLETE') else (
            '#f6b26b' if sg.startswith('PARTIAL') else (
                YELLOW if present7 else None))
        scol = 'color:#2e7d32' if sg.startswith('COMPLETE') else (
            'color:#ed7d31' if sg.startswith('PARTIAL') else '')
        ws.cell(r, 1, value=name)
        done_style(ws, r, 1, wrap=True, bold=True, bg=sbg)
        link_cell(ws, r, 1, name, PLATFORM + '/#' + sid, bold=True, bg=sbg,
                  wrap=True)
        ws.cell(r, 2, value=v(OVR, SH, sid, 'Priority', x['prio']))
        done_style(ws, r, 2, center=True, bold=True)
        for ci, key, base_key in [(3, 'RFC BHMPS', 'bhm'),
                                  (4, 'RFC EIT', 'eit'),
                                  (5, 'Baseline', 'base'),
                                  (6, 'Recovery', 'rec')]:
            ws.cell(r, ci,
                    value=v(OVR, SH, sid, key, x[base_key]))
            done_style(ws, r, ci, center=True, size=10)
        ws.cell(r, 7, value=v(OVR, SH, sid, 'SIGNED', x['signed'] or '-'))
        done_style(ws, r, 7, center=True, bold=True,
                   color=GREEN if sg.startswith('COMPLETE') else
                   (ORANGE if sg.startswith('PARTIAL') else None))
        ws.cell(r, 8, value=v(OVR, SH, sid, 'Milestone', x['mile']))
        done_style(ws, r, 8, size=10)
        ws.cell(r, 9, value='-' if tot == '' else f"{tot}%")
        done_style(ws, r, 9, center=True, bold=True,
                   color=pct_color(int(tot)) if str(tot).isdigit() else GRAY)
        for k, dd in enumerate(x['d']):
            base = grp_start + k
            bg = DC[RFK[k]]
            ws.cell(r, base,
                    value=v(OVR, SH, sid, RFK[k] + ' TOTAL', dd['t']))
            done_style(ws, r, base, center=True, bg=bg, size=10)
            ws.cell(r, base + 1,
                    value=v(OVR, SH, sid, RFK[k] + ' CLOSED', dd['c']))
            done_style(ws, r, base + 1, center=True, bg=bg, size=10)
            p = round(dd['c'] / dd['t'] * 100) if dd['t'] > 0 else (
                100 if dd['c'] > 0 else None)
            ws.cell(r, base + 2,
                    value='-' if p is None else f"{p}%")
            done_style(ws, r, base + 2, center=True, bg=bg, size=10, bold=True,
                       color=pct_color(p) if p is not None else '#bbbbbb')
        col_it = len(fixed) + 25       # ITRs = fixed9 + 24 = col33
        ws.cell(r, col_it, value=v(OVR, SH, sid, 'ITRs', x['it']))
        done_style(ws, r, col_it, center=True, bold=True)
        ws.cell(r, col_it + 1, value=v(OVR, SH, sid, 'CLOSED', x['ic']))
        done_style(ws, r, col_it + 1, center=True)
        ws.cell(r, col_it + 2, value=v(OVR, SH, sid, 'BALANCE', x['ib']))
        done_style(ws, r, col_it + 2, center=True)
        ws.cell(r, col_it + 3, value='-' if x['itrp'] == '' else f"{x['itrp']}%")
        itrpc = pct_color(int(x['itrp'])) if str(x['itrp']).isdigit() else GRAY
        done_style(ws, r, col_it + 3, center=True, bold=True, color=WHITE,
                   bg=itrpc)
        col_cb = col_it + 4            # CPP / EIT / EACP
        for off, key, base in [(0, 'CPP-1', x['b1']), (1, 'EIT', x['b2']),
                               (2, 'EACOP', x['b3'])]:
            wv = v(OVR, SH, sid, key, base)
            ws.cell(r, col_cb + off, value=wv)
            blk = (base or 0) > 0
            done_style(ws, r, col_cb + off, center=True, bold=blk,
                       color=WHITE if blk else None, bg=BLACK if blk else None)
        col_rem = col_cb + 3
        for off, key, bkey in [(0, 'REMARK EACOP', 're1'),
                               (1, 'REMARK CPP-EIT', 're2'),
                               (2, 'REMARK CPP-1', 're3')]:
            ws.cell(r, col_rem + off,
                    value=v(OVR, SH, sid, key, x[bkey]))
            done_style(ws, r, col_rem + off, wrap=True, size=9, bg=REMARK)
        ws.cell(r, col_rem + 3,
                value=v(OVR, SH, sid, 'STATUS', x['wd']))
        done_style(ws, r, col_rem + 3, center=True)
        col_pcol = col_rem + 4
        pcc = PCOL.get(SH, {}).get(sid) or ''
        ws.cell(r, col_pcol)
        done_style(ws, r, col_pcol, center=True, bg=PCOL_NUM[0][1] if False else (pcc if pcc else None))
        col_notes = col_pcol + 1
        ws.cell(r, col_notes, value=(NOTES.get(SH, {}) or {}).get(sid, ''))
        done_style(ws, r, col_notes, wrap=True, size=9)
        note_fonts.append((r, col_notes, (NOTES.get(SH, {}) or {}).get(sid, '')))
        r += 1
    w.widths(ws, widths)
    ws.freeze_panes = 'D4'
    w.finalize(ws, r - 1, 3, 46)

    # ---------------- BLOCKERS ----------------
    ws = w.sheet('BLOCKERS')
    SH = 'RFC PROGRESS'
    d = sorted([x for x in RFC],
               key=lambda x: -((x['b1'] or 0) + (x['b2'] or 0) + (x['b3'] or 0)))
    heads = ['SUB SYSTEM', 'PRIO', 'TOTAL %', 'CPP-1', 'EIT', 'EACOP',
             'REMARK EACOP', 'REMARK CPP-EIT', 'REMARK CPP-1', 'WALKDOWN',
             '\u25cf', 'MY REMARKS']
    w.title(ws, len(heads), 'BLOCKERS - sorted by blocking count')
    w.head(ws, 2, heads, height=26)
    w.widths(ws, [40, 6, 9, 7, 7, 8, 20, 20, 20, 14, 4, 30])
    r = 3
    for x in d:
        sid = x['sid']
        pcc = PCOL.get(SH, {}).get(sid) or ''
        ws.cell(r, 1, value=x['name'])
        done_style(ws, r, 1, wrap=True, bold=True, bg=pcc or None)
        link_cell(ws, r, 1, x['name'], PLATFORM + '/#' + sid, bold=True,
                  bg=pcc or None, wrap=True)
        ws.cell(r, 2, value=x['prio'])
        done_style(ws, r, 2, center=True, bg=pcc or None)
        ws.cell(r, 3, value='-' if x['tot'] == '' else f"{x['tot']}%")
        done_style(ws, r, 3, center=True, bg=pcc or None,
                   color=pct_color(int(x['tot'])) if str(x['tot']).isdigit()
                   else GRAY)
        for off, key, base in [(0, 'CPP-1', x['b1']), (1, 'EIT', x['b2']),
                               (2, 'EACOP', x['b3'])]:
            ws.cell(r, 4 + off, value=v(OVR, SH, sid, key, base))
            done_style(ws, r, 4 + off, center=True, bg=pcc or None)
        for off, key, bkey in [(0, 'REMARK EACOP', 're1'),
                               (1, 'REMARK CPP-EIT', 're2'),
                               (2, 'REMARK CPP-1', 're3')]:
            ws.cell(r, 7 + off, value=v(OVR, SH, sid, key, x[bkey]))
            done_style(ws, r, 7 + off, wrap=True, size=9, bg=REMARK)
        ws.cell(r, 10, value=v(OVR, SH, sid, 'STATUS', x['wd']))
        done_style(ws, r, 10, center=True, bg=pcc or None)
        ws.cell(r, 11)
        done_style(ws, r, 11, center=True, bg=pcc or None)
        ws.cell(r, 12, value=(NOTES.get(SH, {}) or {}).get(sid, ''))
        done_style(ws, r, 12, wrap=True, size=9)
        r += 1
    w.finalize(ws, r - 1, 2, len(heads))

    # ---------------- PUNCH LIST ----------------
    ws = w.sheet('PUNCH LIST')
    SH = 'PUNCH LIST'
    heads = ['ID', 'TAG', 'CAT', 'DISC', 'DESCRIPTION', 'STATUS', 'CLOSE',
             'SUB SYSTEM', '\u25cf', 'NOTES']
    w.title(ws, len(heads), f'PUNCH LIST - {len(PUNCH)} item(s)')
    w.head(ws, 2, heads, height=24)
    w.widths(ws, [13, 16, 6, 5, 55, 14, 14, 30, 4, 30])
    r = 3
    for rec in PUNCH:
        pid = str(rec[0])
        bs = PCOL.get(SH, {}).get(pid)
        ws.cell(r, 1, value=pid)
        done_style(ws, r, 1, bold=True, bg=bs or None)
        ws.cell(r, 2, value=v(OVR, SH, pid, 'TAG', rec[1]))
        done_style(ws, r, 2, bg=bs or None)
        cat = v(OVR, SH, pid, 'CAT', rec[2])
        ws.cell(r, 3, value=cat)
        cb = CAT_FILL.get(cat)
        done_style(ws, r, 3, center=True, bold=True, bg=cb or None)
        ws.cell(r, 4, value=v(OVR, SH, pid, 'DISC', rec[3]))
        done_style(ws, r, 4, center=True, bg=bs or None)
        ws.cell(r, 5, value=v(OVR, SH, pid, 'DESCRIPTION', rec[4]))
        done_style(ws, r, 5, wrap=True)
        st = v(OVR, SH, pid, 'STATUS', rec[5])
        ws.cell(r, 6, value=st)
        if st.strip().lower() in ('closed', 'completed'):
            done_style(ws, r, 6, bold=True, color=GREEN)
        elif re.search(r'originated|to\s*be', st, re.I):
            done_style(ws, r, 6, color='#b35900')
        else:
            done_style(ws, r, 6)
        ws.cell(r, 7, value=v(OVR, SH, pid, 'CLOSING DATE', rec[6]))
        done_style(ws, r, 7, center=True, bg=bs or None)
        ws.cell(r, 8, value=rec[8])
        done_style(ws, r, 8, wrap=True, bg=bs or None)
        link_cell(ws, r, 8, rec[8], PLATFORM + '/#' + str(rec[9]),
                  bg=bs or None, wrap=True)
        ws.cell(r, 9)
        done_style(ws, r, 9, center=True, bg=bs or None)
        ws.cell(r, 10, value=(NOTES.get(SH, {}) or {}).get(pid, ''))
        done_style(ws, r, 10, wrap=True, size=9)
        r += 1
    w.finalize(ws, r - 1, 2, len(heads))

    # ---------------- ITR LIST ----------------
    ws = w.sheet('ITR LIST')
    SH = 'ITR LIST'
    heads = ['ID', 'TAG', 'DISC', 'TASK TYPE', 'DESCRIPTION', 'STATE',
             'CLOSE', 'SUB SYSTEM', '\u25cf', 'NOTES']
    w.title(ws, len(heads), f'ITR LIST - {len(ITR)} task(s)')
    w.head(ws, 2, heads, height=24)
    w.widths(ws, [15, 16, 5, 12, 55, 22, 14, 30, 4, 30])
    r = 3
    for rec in ITR:
        tid = str(rec[0])
        bs = PCOL.get(SH, {}).get(tid)
        ws.cell(r, 1, value=tid)
        done_style(ws, r, 1, bold=True, bg=bs or None)
        ws.cell(r, 2, value=v(OVR, SH, tid, 'TAG', rec[1]))
        done_style(ws, r, 2, bg=bs or None)
        ws.cell(r, 3, value=v(OVR, SH, tid, 'DISC', rec[2]))
        done_style(ws, r, 3, center=True, bg=bs or None)
        tt = v(OVR, SH, tid, 'TASK TYPE', '') or rec[3]
        ws.cell(r, 4, value=tt)
        done_style(ws, r, 4, bg=bs or None)
        ws.cell(r, 5, value=v(OVR, SH, tid, 'ASSET DESCRIPTION', rec[4]))
        done_style(ws, r, 5, wrap=True)
        st = v(OVR, SH, tid, 'STATE', rec[5])
        ws.cell(r, 6, value=st)
        if done_state(st):
            done_style(ws, r, 6, bold=True, color=GREEN)
        elif re.search(r'originated|to\s*be', st, re.I):
            done_style(ws, r, 6, color='#b35900')
        else:
            done_style(ws, r, 6)
        ws.cell(r, 7, value=v(OVR, SH, tid, 'CLOSING DATE', rec[6]))
        done_style(ws, r, 7, center=True, bg=bs or None)
        ws.cell(r, 8, value=rec[8])
        done_style(ws, r, 8, wrap=True, bg=bs or None)
        link_cell(ws, r, 8, rec[8], PLATFORM + '/#' + str(rec[9]),
                  bg=bs or None, wrap=True)
        ws.cell(r, 9)
        done_style(ws, r, 9, center=True, bg=bs or None)
        ws.cell(r, 10, value=(NOTES.get(SH, {}) or {}).get(tid, ''))
        done_style(ws, r, 10, wrap=True, size=9)
        r += 1
    w.finalize(ws, r - 1, 2, len(heads))

    # ---------------- SUB SYSTEM ----------------
    ws = w.sheet('SUB SYSTEM')
    heads = ['SID', 'SUB SYSTEM', 'PUNCH T', 'PUNCH CL', 'PUNCH %', 'ITR T',
             'ITR CL', 'ITR %', 'E CPP CL/T', 'I CPP CL/T', 'T CPP CL/T',
             'TOTAL %', 'RFC SIGNED']
    w.title(ws, len(heads), f'SUB SYSTEM - {len(SUBS)} subsystem(s)')
    w.head(ws, 2, heads, height=26)
    w.widths(ws, [10, 42, 9, 9, 9, 9, 9, 9, 13, 13, 13, 9, 14])
    r = 3
    for s in SUBS:
        sid = s['sid']
        P = [p for p in PUNCH if str(p[9]) == sid]
        I = [t for t in ITR if str(t[9]) == sid]
        pc = sum(1 for p in P
                 if str(p[5]).strip().lower() in ('closed', 'completed'))
        ic = sum(1 for t in I if done_state(t[5]))
        pp = round(pc / len(P) * 100) if P else 0
        ip = round(ic / len(I) * 100) if I else 0
        E = [t for t in I if t[2] == 'E' and t[7] == 'CPP AGI']
        IN = [t for t in I if t[2] == 'I' and t[7] == 'CPP AGI']
        T = [t for t in I if t[2] == 'T' and t[7] == 'CPP AGI']
        def cl(pool):
            c = sum(1 for t in pool if done_state(t[5]))
            return f"{c}/{len(pool)}"
        x = next((x for x in RFC if x['sid'] == sid), None)
        ws.cell(r, 1, value=sid)
        done_style(ws, r, 1, center=True, bold=True)
        ws.cell(r, 2, value=s['name'])
        done_style(ws, r, 2, wrap=True)
        link_cell(ws, r, 2, s['name'], PLATFORM + '/#' + sid, wrap=True)
        ws.cell(r, 3, value=len(P))
        done_style(ws, r, 3, center=True)
        ws.cell(r, 4, value=pc)
        done_style(ws, r, 4, center=True)
        ws.cell(r, 5, value=f"{pp}%")
        done_style(ws, r, 5, center=True, bold=True,
                   color=pct_color(pp) if P else GRAY)
        ws.cell(r, 6, value=len(I))
        done_style(ws, r, 6, center=True)
        ws.cell(r, 7, value=ic)
        done_style(ws, r, 7, center=True)
        ws.cell(r, 8, value=f"{ip}%")
        done_style(ws, r, 8, center=True, bold=True,
                   color=pct_color(ip) if I else GRAY)
        ws.cell(r, 9, value=cl(E))
        done_style(ws, r, 9, center=True)
        ws.cell(r, 10, value=cl(IN))
        done_style(ws, r, 10, center=True)
        ws.cell(r, 11, value=cl(T))
        done_style(ws, r, 11, center=True)
        tv = x['tot'] if x else ''
        ws.cell(r, 12, value='-' if tv == '' else f"{tv}%")
        done_style(ws, r, 12, center=True, bold=True,
                   color=pct_color(int(tv)) if str(tv).isdigit() else GRAY)
        sg = ''
        if x:
            sg = v(OVR, 'RFC PROGRESS', sid, 'SIGNED', x['signed'])
        ws.cell(r, 13,
                value=v(OVR, 'RFC PROGRESS', sid, 'SIGNED', x['signed'] or '-')
                if x else '-')
        sgu = str(sg or '').upper()
        done_style(ws, r, 13, center=True, bold=True,
                   color=GREEN if sgu.startswith('COMPLETE') else
                   (ORANGE if sgu.startswith('PARTIAL') else None))
        r += 1
    w.finalize(ws, r - 1, 2, len(heads))

    # ---------------- DONE ----------------
    ws = w.sheet('DONE')
    SH = 'RFC PROGRESS'
    heads = ['#', 'SUB SYSTEM', 'TOTAL %', 'MILESTONE', 'RFC SIGNED',
             'WALKDOWN STATUS', 'MY REMARKS']
    w.title(ws, len(heads), 'DONE - subsystems at 100%')
    w.head(ws, 2, heads, height=24)
    w.widths(ws, [5, 46, 9, 16, 14, 16, 34])
    r = 3
    d = [x for x in RFC
         if (x['tot'] != '' and x['tot'] >= 100)
         or str(x['signed']).upper() == 'COMPLETE'
         or str(x['signed']).upper() == 'PARTIAL']
    for i, x in enumerate(d, 1):
        sid = x['sid']
        tot_ok = x['tot'] != '' and x['tot'] >= 100
        ws.cell(r, 1, value=i)
        done_style(ws, r, 1, center=True, bold=True,
                   bg=GREENC if tot_ok else YELLOW)
        ws.cell(r, 2, value=x['name'])
        done_style(ws, r, 2, wrap=True)
        link_cell(ws, r, 2, x['name'], PLATFORM + '/#' + sid, wrap=True)
        ws.cell(r, 3, value='-' if x['tot'] == '' else f"{x['tot']}%")
        done_style(ws, r, 3, center=True, bold=True)
        ws.cell(r, 4, value=v(OVR, SH, sid, 'Milestone', x['mile']))
        done_style(ws, r, 4, size=10)
        sgu7 = str(x['signed']).upper()
        ws.cell(r, 5, value=v(OVR, SH, sid, 'SIGNED', x['signed'] or '-'))
        done_style(ws, r, 5, center=True, bold=True,
                   color=GREEN if sgu7.startswith('COMPLETE') else
                   (ORANGE if sgu7.startswith('PARTIAL') else None))
        ws.cell(r, 6, value=v(OVR, SH, sid, 'STATUS', x['wd']))
        done_style(ws, r, 6, center=True)
        ws.cell(r, 7, value=(NOTES.get(SH, {}) or {}).get(sid, ''))
        done_style(ws, r, 7, wrap=True, size=9)
        r += 1
    w.finalize(ws, r - 1, 2, len(heads))

    # ---------------- CABLES ----------------
    ws = w.sheet('CABLES')
    heads = ['#', 'DISC', 'CABLE TAG', 'FROM TAG', 'FROM DESC', 'FROM LOC',
             'TO TAG', 'TO DESC', 'TO LOC', 'PULLING', 'RFI#', 'EIT DT',
             'RFC DT', 'PUNCH']
    w.title(ws, len(heads), f'CABLES - {len(CABLES)} cable(s)')
    w.head(ws, 2, heads, height=24)
    w.widths(ws, [5, 12, 20, 16, 30, 14, 16, 30, 14, 12, 10, 12, 12, 20])
    r = 3
    for i, rec in enumerate(CABLES, 1):
        ws.cell(r, 1, value=i)
        done_style(ws, r, 1, center=True)
        disc = rec[1]
        ws.cell(r, 2, value=disc)
        done_style(ws, r, 2, bg=CAB_DC.get(disc), bold=True, center=True)
        ws.cell(r, 3, value=rec[2])
        done_style(ws, r, 3, bold=True)
        ws.cell(r, 4, value=rec[3])
        done_style(ws, r, 4)
        ws.cell(r, 5, value=rec[4])
        done_style(ws, r, 5)
        ws.cell(r, 6, value=rec[5])
        done_style(ws, r, 6)
        ws.cell(r, 7, value=rec[6])
        done_style(ws, r, 7)
        ws.cell(r, 8, value=rec[7])
        done_style(ws, r, 8)
        ws.cell(r, 9, value=rec[8])
        done_style(ws, r, 9)
        ws.cell(r, 10, value=rec[9])
        done_style(ws, r, 10, center=True)
        ws.cell(r, 11, value=rec[10])
        done_style(ws, r, 11)
        ws.cell(r, 12, value=rec[11])
        done_style(ws, r, 12, center=True)
        ws.cell(r, 13, value=rec[12])
        done_style(ws, r, 13, center=True)
        ws.cell(r, 14, value=rec[13])
        done_style(ws, r, 14, wrap=True)
        r += 1
    w.finalize(ws, r - 1, 2, len(heads))

    # ---------------- DASHBOARD (KPIs) ----------------
    ws = w.sheet('DASHBOARD')
    itrT = len(ITR)
    itrCl = sum(1 for t in ITR if done_state(t[5]))
    itrOp = itrT - itrCl
    itrPct = round(itrCl / itrT * 100) if itrT else 0
    puT = len(PUNCH)
    puCl = sum(1 for p in PUNCH if str(p[5]).strip() == 'Closed')
    puOp = puT - puCl
    puPct = round(puCl / puT * 100) if puT else 0
    comp100 = sum(1 for x in RFC if isinstance(x['tot'], (int, float))
                  and x['tot'] >= 100)
    kpis = [
        ('ITR TOTAL', itrT, '#0b2239'),
        ('ITR CLOSED', itrCl, '#256a29'),
        ('ITR %', f'{itrPct}%', '#00564c'),
        ('PUNCH TOTAL', puT, '#3a62ad'),
        ('PUNCH CLOSED', puCl, '#256a29'),
        ('PUNCH OPEN', puOp, '#c2610c'),
        ('MILESTONES', len(MILES), '#0b2239'),
        ('100% COMPLETE', comp100, '#256a29'),
    ]
    w.title(ws, 10, 'DASHBOARD - live summary (same as page 1)')
    for i, (lab, val, col) in enumerate(kpis, 1):
        vc = ws.cell(2, i, value=val if isinstance(val, int)
                     else val)
        vc.font = Font(name=FONT, size=22, bold=True, color=a(WHITE))
        vc.alignment = Alignment(horizontal='center', vertical='center')
        vc.fill = fill(col)
        lc = ws.cell(3, i, value=lab)
        lc.font = Font(name=FONT, size=9, bold=True, color=a(WHITE))
        lc.alignment = Alignment(horizontal='center', vertical='center')
        lc.fill = fill(col)
        ws.column_dimensions[get_column_letter(i)].width = 15
    for i in range(1, 9):
        ws.cell(2, i).border = BORDER
        ws.cell(3, i).border = BORDER
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 18
    nav_sheets = ['1 \u00b7 DASHBOARD', '2 \u00b7 RFC PROGRESS',
                  '3 \u00b7 BLOCKING \u2013 REMARKS', '4 \u00b7 PUNCH LIST',
                  '5 \u00b7 ITR LIST', '6 \u00b7 SUBSYSTEM',
                  '7 \u00b7 COMPLETE', '8 \u00b7 EIT CABLES']
    ws.cell(4, 1, value='TABS \u25b6')
    ws.cell(4, 1).font = Font(name=FONT, size=10, bold=True, color=a(NAVY))
    ws.cell(4, 1).alignment = Alignment(vertical='center', indent=1)
    ws.cell(4, 1).border = BORDER
    for i2, nm in enumerate(nav_sheets, 2):
        link_cell(ws, 4, i2, nm, "#'" + nm + "'!A1", bold=True, size=10,
                  center=True)
    ws.row_dimensions[4].height = 20
    ws.freeze_panes = ws.cell(5, 1).coordinate
    def banner(rr, text, color='#1f4e79'):
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr,
                       end_column=10)
        for cc_ in range(1, 11):
            ws.cell(rr, cc_).fill = fill(color)
            ws.cell(rr, cc_).border = BORDER
        cell = ws.cell(rr, 1, value=text)
        cell.font = Font(name=FONT, size=11, bold=True, color=a(WHITE))
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[rr].height = 20

    # Section 1 - ITR PROGRESS BY DISCIPLINE (website panel)
    banner(5, 'ITR PROGRESS BY DISCIPLINE')
    w.head(ws, 6, ['GROUP', 'DISCIPLINE', 'TOTAL', 'CLOSED', '%', 'OPEN'],
           height=22)
    r = 8
    for rec in ITRT:
        tot_row = str(rec[1]).lower().find('total') >= 0
        p = None if rec[4] in (None, '') else round(rec[4] * 100)
        ws.cell(r, 1, value=rec[0])
        ws.cell(r, 2, value=rec[1])
        ws.cell(r, 3, value=rec[2])
        ws.cell(r, 4, value=rec[3])
        ws.cell(r, 5, value='' if p is None else f'{p}%')
        ws.cell(r, 6, value=rec[5])
        for c in range(1, 7):
            done_style(ws, r, c, bold=tot_row,
                       bg='#ffe9c9' if str(rec[0]).lower() == 'grand'
                       else None,
                       center=True if c != 2 else False)
        if p is not None:
            ws.cell(r, 5).font = Font(name=FONT, size=10, bold=True,
                                      color=a(WHITE))
            ws.cell(r, 5).fill = fill(pct_color(p))
        r += 1

    # Section 2 - PUNCH SUMMARY (A / B / C) (website panel)
    r += 2
    banner(r, 'PUNCH SUMMARY (A / B / C)')
    w.head(ws, r + 1, ['DISCIPLINE', 'A', 'B', 'C'],
           spans={2: 3, 3: 3, 4: 3},
           fills={2: '#4472c4', 3: '#ed7d31', 4: '#57a05b'}, height=22)
    w.head(ws, r + 2, ['', 'T', 'CL', 'OP', 'T', 'CL', 'OP', 'T', 'CL', 'OP'],
           height=16)
    r += 3
    for rec in PUNT:
        grand = str(rec[0]).lower().startswith('grand')
        ws.cell(r, 1, value=rec[0])
        for i in range(1, 10):
            ws.cell(r, 1 + i, value=rec[i])
        for c in range(1, 11):
            done_style(ws, r, c, bold=grand,
                       center=True if c > 1 else False,
                       bg='#ffe9c9' if grand else None)
        r += 1

    # Section 3 - MILESTONE & DIS (website panel)
    r += 2
    banner(r, f'MILESTONE & DIS - {len(MILES)} ROWS')
    w.head(ws, r + 1, ['SUB', 'DESCRIPTION', 'MILESTONE', 'MONTH', 'MANTRAC',
                       'PRIORITY', 'PG', 'RFSU', 'MP DATE', 'EIT DATE'],
           height=22)
    r += 2
    for x in MILES:
        vals = [x.get('sub', ''), x.get('desc', ''), x.get('milestone', ''),
                x.get('month', ''), x.get('mantrac', ''),
                x.get('priority', ''), x.get('pg', ''), x.get('rfsu', ''),
                x.get('mp_date', ''), x.get('eit_date', '')]
        for c in range(1, 11):
            ws.cell(r, c, value=vals[c - 1])
            done_style(ws, r, c, wrap=True if c == 2 else False)
        r += 1
    w.widths(ws, [14, 38, 22, 9, 13, 9, 8, 8, 11, 11])
    ws.freeze_panes = 'A4'

    # (MILESTONES / ITR TRACKER / PUNCH SUMMARY are inside sheet 1, like the
    # website DASHBOARD page)

    # ---------------- READ ME ----------------
    ws = w.sheet('READ ME')
    rows = [
        ('PS5 COMPLETION PLATFORM - EXCEL (all sheets match the online platform)',
         True, 14, None, None),
        ('', False, 11, None, None),
        ('HOW IT WORKS', True, 12, NAVY, WHITE),
        ('This Excel is a LIVE copy of the platform website. Any edit you type here', False, 11, None, None),
        ('and SAVE via the desktop icon is pushed to the platform ONLINE and is seen', False, 11, None, None),
        ('by every laptop that opens it. The 3 accounts can all use it from any laptop.', False, 11, None, None),
        ('', False, 11, None, None),
        ('STEPS (any laptop, any of the 3 accounts)', True, 12, NAVY, WHITE),
        ('1. Double-click the desktop icon "PS5 PLATFORM EXCEL" -> the file opens.', False, 11, None, None),
        ('2. Type your edits directly in the colored cells (same as on the website).', False, 11, None, None),
        ('3. Save edits WITHOUT leaving Excel (Ctrl+S), then close Excel.', False, 11, None, None),
        ('4. Double-click the icon again -> your edits are pushed ONLINE, the latest', False, 11, None, None),
        ('   state from the platform is pulled, and the file updates + reopens.', False, 11, None, None),
        ('', False, 11, None, None),
        ('SET UP THE ONLINE TOKEN (once, on each laptop)', True, 12, NAVY, WHITE),
        ('On the first run the icon will ask you to paste your GitHub access token.', False, 11, None, None),
        ('It is kept ONLY in this folder (platform_token.txt) for that laptop.', False, 11, None, None),
        ('', False, 11, None, None),
        ('SHEETS (all 8 platform pages - same names, colours & formatting)', True, 12, NAVY, WHITE),
        ('1 \u00b7 DASHBOARD            = page 1 - KPI cards + ITR progress + Punch A/B/C + milestones', False, 11, None, None),
        ('2 \u00b7 RFC PROGRESS         = page 2 - matrix (editable), disciplines B E H I M P S T', False, 11, None, None),
        ('3 \u00b7 BLOCKING \u2013 REMARKS = page 3 - blocking points & remarks (CPP-1 / EIT / EACOP)', False, 11, None, None),
        ('4 \u00b7 PUNCH LIST           = page 4 - detailed punch (CAT A/red, B/yellow, C/green)', False, 11, None, None),
        ('5 \u00b7 ITR LIST             = page 5 - detailed ITR tasks', False, 11, None, None),
        ('6 \u00b7 SUBSYSTEM            = page 6 - per-subsystem punch / ITR / E-I-T summaries', False, 11, None, None),
        ('7 \u00b7 COMPLETE             = page 7 - subsystems at 100%', False, 11, None, None),
        ('8 \u00b7 EIT CABLES           = page 8 - electrical / instrument / telecom / trace-heating cables', False, 11, None, None),
        ('', False, 11, None, None),
        ('LINKS: every subsystem NAME is a live link (blue underlined) - click it', False, 11, None, None),
        ('to open that subsystem on the platform. Row 4 of DASHBOARD has TABS links', False, 11, None, None),
        ('to jump between sheets.', False, 11, None, None),
        ('TIP: you can also edit online on the website - any edit there appears here',
         False, 10, None, None),
        ('after the next icon refresh, and vice versa.', False, 10, None, None),
    ]
    ws.column_dimensions['A'].width = 110
    r = 1
    for txt, bold, size, bg, fg in rows:
        cell = ws.cell(r, 1, value=txt)
        cell.font = Font(name=FONT, size=size, bold=bold,
                         color='FF0B2239' if not fg else a(WHITE))
        if bg:
            for col in range(1, 6):
                ws.cell(r, col).fill = fill(bg)
        r += 1

    # ---------------- SIDMAP (hidden) ----------------
    wb = w.wb
    sm = wb.create_sheet('SIDMAP')
    sm.sheet_state = 'hidden'
    for x in RFC:
        sm.append([x['name'], x['sid']])

    # ---- final: sheets named exactly like the 8 platform pages, in order ----
    REN = [
        ('READ ME', '0 \u00b7 READ ME'),
        ('DASHBOARD', '1 \u00b7 DASHBOARD'),
        ('RFC PROGRESS', '2 \u00b7 RFC PROGRESS'),
        ('BLOCKERS', '3 \u00b7 BLOCKING \u2013 REMARKS'),
        ('PUNCH LIST', '4 \u00b7 PUNCH LIST'),
        ('ITR LIST', '5 \u00b7 ITR LIST'),
        ('SUB SYSTEM', '6 \u00b7 SUBSYSTEM'),
        ('DONE', '7 \u00b7 COMPLETE'),
        ('CABLES', '8 \u00b7 EIT CABLES'),
    ]
    for old, new in REN:
        if old in wb.sheetnames:
            wb[old].title = new
    order = [new for _, new in REN if new in wb.sheetnames]
    for extra in wb.sheetnames:
        if extra not in order:
            order.append(extra)
    wb._sheets = [wb[nm] for nm in order]
    if '0 \u00b7 READ ME' in wb.sheetnames:
        wb.active = wb.sheetnames.index('0 \u00b7 READ ME')

    return w.wb


# ------------------------------------------------------------------- page 9 --
def build_eit9_workbook(ITR, RFC, OVR, NOTES, PCOL):
    """Standalone workbook for platform page 9 - ITR OPEN E/I/T (CPP AGI)."""
    w = WB()
    ws = w.sheet('9 \u00b7 ITR OPEN EIT')
    pool = [t for t in ITR if t[2] in ('E', 'I', 'T')
            and str(t[7]).strip() == 'CPP AGI']
    open_pool = [t for t in pool if not done_state(t[5])]
    by = {d: {'t': 0, 'c': 0} for d in ('E', 'I', 'T')}
    for t in pool:
        by[t[2]]['t'] += 1
        if done_state(t[5]):
            by[t[2]]['c'] += 1
    colors = {'E': '#1a5276', 'I': '#b9770e', 'T': '#1b4f72'}
    disc_names = {'E': 'E - ELECTRICAL', 'I': 'I - INSTRUMENT',
                  'T': 'T - TELECOM'}
    hearts = []
    for d in ('E', 'I', 'T'):
        st = by[d]
        hearts.append((disc_names[d], st['t'], st['c'], colors[d]))
    w.title(ws, 9, '9 \u00b7 ITR OPEN E/I/T  |  CPP AGI open tasks')
    for i, (lab, tot2, cl2, col) in enumerate(hearts, 1):
        pct2 = round(cl2 / tot2 * 100) if tot2 else 0
        vc = ws.cell(2, i, value=f'{pct2}%')
        vc.font = Font(name=FONT, size=18, bold=True, color=a(WHITE))
        vc.alignment = Alignment(horizontal='center', vertical='center')
        vc.fill = fill(col)
        lc = ws.cell(3, i, value=f'{lab}  |  {cl2}/{tot2} closed')
        lc.font = Font(name=FONT, size=8, bold=True, color=a(WHITE))
        lc.alignment = Alignment(horizontal='center', vertical='center',
                                 wrap_text=True)
        lc.fill = fill(col)
        ws.column_dimensions[get_column_letter(i)].width = 15
        ws.cell(2, i).border = BORDER
        ws.cell(3, i).border = BORDER
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 22

    def banner(rr, text, color='#1f4e79'):
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=9)
        for cc_ in range(1, 10):
            ws.cell(rr, cc_).fill = fill(color)
            ws.cell(rr, cc_).border = BORDER
        cell = ws.cell(rr, 1, value=text)
        cell.font = Font(name=FONT, size=11, bold=True, color=a(WHITE))
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[rr].height = 20

    banner(5, f'ITR OPEN E/I/T (CPP AGI) - {len(open_pool)} OPEN TASK(S) '
              f'OUT OF {len(pool)} TOTAL', color=colors['E'])
    heads = ['#', 'DISC', 'TAG', 'ASSET DESCRIPTION', 'STATE', 'CLOSE DATE',
             'SUB SYSTEM', 'TASK TYPE', '\u25cf']
    w.head(ws, 6, heads, height=24)
    w.widths(ws, [5, 6, 16, 55, 22, 14, 32, 14, 4])
    r = 7
    for i, t in enumerate(open_pool, 1):
        tid = str(t[0])
        bs = PCOL.get('ITR LIST', {}).get(tid) or ''
        ws.cell(r, 1, value=i)
        done_style(ws, r, 1, center=True, bold=True, bg=bs or None)
        dcol = colors.get(t[2], '#555')
        ws.cell(r, 2, value=t[2])
        done_style(ws, r, 2, center=True, bold=True, bg=dcol,
                   color=WHITE)
        ws.cell(r, 3, value=t[1])
        done_style(ws, r, 3, bg=bs or None)
        ws.cell(r, 4, value=t[4])
        done_style(ws, r, 4, wrap=True)
        ws.cell(r, 5, value=v(OVR, 'ITR LIST', tid, 'STATE', t[5]))
        done_style(ws, r, 5, center=True, size=10)
        ws.cell(r, 6, value=t[6])
        done_style(ws, r, 6, center=True, bg=bs or None)
        ws.cell(r, 7, value=t[8])
        done_style(ws, r, 7, wrap=True, bg=bs or None)
        link_cell(ws, r, 7, t[8], PLATFORM + '/#' + str(t[9]),
                  bg=bs or None, wrap=True)
        ws.cell(r, 8, value=t[3])
        done_style(ws, r, 8, bg=bs or None)
        ws.cell(r, 9)
        done_style(ws, r, 9, center=True, bg=bs or None)
        r += 1
    w.finalize(ws, r - 1, 6, len(heads))
    return w.wb


# -------------------------------------------------------------------- pages --
PAGE_FILES = [
    ('1', '1 \u00b7 DASHBOARD', '1 - DASHBOARD.xlsx'),
    ('2', '2 \u00b7 RFC PROGRESS', '2 - RFC PROGRESS.xlsx'),
    ('3', '3 \u00b7 BLOCKING \u2013 REMARKS', '3 - BLOCKING-REMARKS.xlsx'),
    ('4', '4 \u00b7 PUNCH LIST', '4 - PUNCH LIST.xlsx'),
    ('5', '5 \u00b7 ITR LIST', '5 - ITR LIST.xlsx'),
    ('6', '6 \u00b7 SUBSYSTEM', '6 - SUBSYSTEM.xlsx'),
    ('7', '7 \u00b7 COMPLETE', '7 - COMPLETE.xlsx'),
    ('8', '8 \u00b7 EIT CABLES', '8 - EIT CABLES.xlsx'),
]


def build_pages(data):
    ITR, RFC, OVR, NOTES, PCOL = data[2], data[3], data[8], data[9], data[10]
    out_dir = os.path.join(HERE, 'PAGES')
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)
    if not os.path.exists(XLSX):
        print('[pages] main workbook missing - run a build first')
        return
    for num, sheet, fname in PAGE_FILES:
        wb = openpyxl.load_workbook(XLSX)
        todel = [n for n in wb.sheetnames if n != sheet and n != 'SIDMAP']
        for n in todel:
            wb.remove(wb[n])
        if sheet in wb.sheetnames:
            wb.active = wb.sheetnames.index(sheet)
        out = os.path.join(out_dir, fname)
        if os.path.exists(out):
            os.remove(out)
        wb.save(out)
        print(f'[pages] saved {fname}  ({sheet})')
    p9 = os.path.join(out_dir, '9 - ITR OPEN EIT.xlsx')
    if os.path.exists(p9):
        os.remove(p9)
    build_eit9_workbook(ITR, RFC, OVR, NOTES, PCOL).save(p9)
    print('[pages] saved 9 - ITR OPEN EIT.xlsx  (9 \u00b7 ITR OPEN E/I/T)')


# ------------------------------------------------------------------ storage --
def read_snapshot():
    try:
        with open(SNAP, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def write_snapshot(data):
    with open(SNAP, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)


def ensure_token():
    if os.path.exists(TOKEN):
        t = open(TOKEN, encoding='utf-8').read().strip()
        if t:
            return t
    print('=' * 52)
    print('FIRST RUN - paste your GitHub access token (from one of the 3 accounts):')
    print('  e.g. ghp_.......................  (has write access to this repo)')
    print('=' * 52)
    print('== NOTE: I will not display what you paste ==')
    tok = input('PASTE TOKEN: ').strip()
    if not tok:
        print('No token given - continuing WITHOUT online push (local only).')
        return ''
    with open(TOKEN, 'w', encoding='utf-8') as f:
        f.write(tok)
    print('Token saved to platform_token.txt (this laptop only).')
    return tok


# -------------------------------------------------------------------- cloud --
def get_state():
    try:
        return json.loads(fetch(RAW_STATE, timeout=30).decode('utf-8'))
    except Exception:
        return {}


def push_state(payload, token):
    """Write platform_state.json via Contents API, retry on 409 (same as web)."""
    if not token:
        print('[push] no token - skipped')
        return False

    def api(method, body):
        req = urllib.request.Request(API_STATE, method=method, data=body)
        req.add_header('Authorization', 'token ' + token)
        req.add_header('Accept', 'application/vnd.github+json')
        try:
            with urllib.request.urlopen(req, timeout=60) as r:
                return r.status, json.loads(r.read().decode('utf-8'))
        except urllib.error.HTTPError as e:
            try:
                return e.code, json.loads(e.read().decode('utf-8'))
            except Exception:
                return e.code, {}

    attempt = 0
    while attempt < 3:
        st, js = api('GET', None)
        if st == 200:
            sha = js.get('sha')
            cur = json.loads(base64.b64decode(js['content']).decode('utf-8'))
            merged = merge_cloud(cur, payload)
            body = json.dumps({
                'message': 'excel sync ' + time.strftime('%H:%M'),
                'content': base64.b64encode(
                    json.dumps(merged, ensure_ascii=False).encode('utf-8')
                ).decode('utf-8'),
                'sha': sha,
            }, ensure_ascii=False).encode('utf-8')
            st2, js2 = api('PUT', body)
            if st2 in (200, 201):
                print('[push] SAVED ONLINE - everyone can now see your edits.')
                return True
            if st2 == 409:
                print(f'[push] conflict ({st2}) - retry {attempt + 1}/3')
                attempt += 1
                time.sleep(1.5)
                continue
            print(f'[push] upload failed: HTTP {st2} - {js2.get("message")}')
            return False
        else:
            print(f'[push] cannot read state: HTTP {st}')
            return False
    print('[push] too many conflicts - keep your edits, try again after 30s')
    return False


def merge_cloud(cloud, mine):
    """mine (local) wins on conflicts - same policy as the web client."""
    cells = (cloud.get('cells') or {}).copy()
    notes = (cloud.get('notes') or {}).copy()
    colors = (cloud.get('colors') or {}).copy()
    for sh, ids in (mine.get('cells') or {}).items():
        for idv, cols in ids.items():
            cells.setdefault(sh, {}).setdefault(idv, {})
            for c, val in cols.items():
                cells[sh][idv][c] = val
    for sh, ids in (mine.get('notes') or {}).items():
        for idv, val in ids.items():
            notes.setdefault(sh, {})[idv] = val
    for sh, ids in (mine.get('colors') or {}).items():
        for idv, val in ids.items():
            colors.setdefault(sh, {})[idv] = val
    return {'cells': cells, 'notes': notes, 'colors': colors}


# -------------------------------------------------------------- push reader --
# Editable columns per sheet (1-based excel columns -> platform col name).
RFC_EDIT = {
    2: 'Priority', 3: 'RFC BHMPS', 4: 'RFC EIT', 5: 'Baseline', 6: 'Recovery',
    7: 'SIGNED', 8: 'Milestone',
}
for k in range(8):
    base = 10 + k * 3
    RFC_EDIT[base] = RFK[k] + ' TOTAL'
    RFC_EDIT[base + 1] = RFK[k] + ' CLOSED'
RFC_EDIT[34] = 'ITRs'
RFC_EDIT[35] = 'CLOSED'
RFC_EDIT[36] = 'BALANCE'
RFC_EDIT[38] = 'CPP-1'
RFC_EDIT[39] = 'EIT'
RFC_EDIT[40] = 'EACOP'
RFC_EDIT[41] = 'REMARK EACOP'
RFC_EDIT[42] = 'REMARK CPP-EIT'
RFC_EDIT[43] = 'REMARK CPP-1'
RFC_EDIT[44] = 'STATUS'          # Walkdown
RFC_EDIT[46] = '__NOTES__'

BLK_EDIT = {4: 'CPP-1', 5: 'EIT', 6: 'EACOP', 7: 'REMARK EACOP',
            8: 'REMARK CPP-EIT', 9: 'REMARK CPP-1', 10: 'STATUS',
            12: '__NOTES__'}
BLK_SID_ONLY = {2, 3}            # name + total% - not pushed

PUNCH_EDIT = {2: 'TAG', 3: 'CAT', 4: 'DISC', 5: 'DESCRIPTION', 6: 'STATUS',
              7: 'CLOSING DATE', 10: '__NOTES__'}
ITR_EDIT = {2: 'TAG', 3: 'DISC', 4: 'TASK TYPE', 5: 'ASSET DESCRIPTION',
            6: 'STATE', 7: 'CLOSING DATE', 10: '__NOTES__'}
# sheet -> (data start row, id col, edit map, cloud sheet, notes->notes, colors col)
SHEETS_FOR_PUSH = [
    ('2 \u00b7 RFC PROGRESS', 4, 1, RFC_EDIT, 'RFC PROGRESS', 46, 45),
    ('3 \u00b7 BLOCKING \u2013 REMARKS', 3, 1, BLK_EDIT, 'RFC PROGRESS', 12, 11),
    ('4 \u00b7 PUNCH LIST', 3, 1, PUNCH_EDIT, 'PUNCH LIST', 10, 9),
    ('5 \u00b7 ITR LIST', 3, 1, ITR_EDIT, 'ITR LIST', 10, 9),
]
SIDMAP_SHEETS = {'2 \u00b7 RFC PROGRESS', '3 \u00b7 BLOCKING \u2013 REMARKS'}


def sidmap_from_wb(wb):
    m = {}
    if 'SIDMAP' in wb.sheetnames:
        ws = wb['SIDMAP']
        for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
            if row[0]:
                m[str(row[0]).strip()] = str(row[1]).strip()
    return m


def read_xlsx_edits(snapshot):
    """Diff current workbook vs snapshot -> patches to push."""
    if not os.path.exists(XLSX):
        return None
    wb = openpyxl.load_workbook(XLSX, data_only=False)
    sidmap = sidmap_from_wb(wb)
    changes = {'cells': {}, 'notes': {}, 'colors': {}}
    for sh_name, start, idc, edmap, cloud_sh, notes_col, color_col in SHEETS_FOR_PUSH:
        if sh_name not in wb.sheetnames:
            continue
        ws = wb[sh_name]
        snap_sh = snapshot.get(sh_name, {})
        for r in range(start, ws.max_row + 1):
            key = ws.cell(r, idc).value
            if not key:
                continue
            key = str(key).strip()
            if not key:
                continue
            idv = (sidmap.get(key, key)
                   if sh_name in SIDMAP_SHEETS else key)
            if not idv:
                continue
            for col, pcol_name in edmap.items():
                cell = ws.cell(r, col)
                val = norm(cell.value)
                was = snap_sh.get(idv, {}).get(pcol_name)
                if was is None:
                    continue            # not previously known -> ignore
                if val == was:
                    continue
                if pcol_name == '__NOTES__':
                    changes['notes'].setdefault(cloud_sh, {})[idv] = val
                else:
                    changes['cells'].setdefault(cloud_sh, {}).setdefault(
                        idv, {})[pcol_name] = val
            # color row
            cc = ws.cell(r, color_col).fill
            current_color = ''
            try:
                if cc and cc.fill_type == 'solid' and cc.start_color and \
                        cc.start_color.rgb and str(cc.start_color.rgb) != '00000000':
                    raw = str(cc.start_color.rgb)
                    current_color = ('#FF' + raw[-6:].upper()
                                     if len(raw) == 8 else '#FF'
                                     + raw.upper())
            except Exception:
                current_color = ''
            was_color = snapshot.get(sh_name, {}).get(idv, {}).get('__COLOR__')
            if was_color is not None and was_color != current_color:
                changes['colors'].setdefault(cloud_sh, {})[idv] = current_color
    wb.close()
    return changes


# --------------------------------------------------------------- pull/save ----
def load_data_local(extra_state):
    """Read index.html + apply extra cloud state to build the workbook."""
    local = os.path.join(HERE, 'index.html')
    use_local = '--local' in sys.argv or '--build' in sys.argv
    html = None
    if use_local and os.path.exists(local):
        html = open(local, encoding='utf-8').read()
        print('[data] using LOCAL index.html')
    else:
        try:
            html = fetch(RAW_INDEX).decode('utf-8')
            print('[data] fetched live index.html from platform')
        except Exception as e:
            if os.path.exists(local):
                html = open(local, encoding='utf-8').read()
                print(f'[data] fetch failed ({e}) - using local index.html')
            else:
                print(f'[data] FATAL: cannot fetch index.html ({e})')
                return None
    ovr_add = (extra_state.get('cells') or {}) if extra_state else {}
    notes_add = (extra_state.get('notes') or {}) if extra_state else {}
    pcol_add = (extra_state.get('colors') or {}) if extra_state else {}
    return parse_index(html, ovr_add, notes_add, pcol_add)


def build_and_save():
    wb = build_workbook(*data)

    def try_save():
        tmp = XLSX + '.tmp'
        wb.save(tmp)
        os.replace(tmp, XLSX)

    for attempt in range(20):
        try:
            try_save()
            return True
        except PermissionError:
            print(f'[save] Excel has the file open - please CLOSE it '
                  f'({attempt + 1}/20)...')
            time.sleep(2)
    print('[save] file stays locked - couldn not update. Close Excel and retry.')
    return False


def publish_xlsx(token):
    """Upload the golden xlsx to the repo so the platform download button
    always serves the latest full workbook (all pages + colours + links)."""
    if not token:
        print('[publish] no token - skipped xlsx upload')
        return False
    if not os.path.exists(XLSX):
        print('[publish] no xlsx yet')
        return False
    api = f'https://api.github.com/repos/{REPO}/contents/' + \
        'PS5%20PLATFORM.xlsx'

    def req(method, body=None):
        rq = urllib.request.Request(api, method=method, data=body)
        rq.add_header('Authorization', 'token ' + token)
        rq.add_header('Accept', 'application/vnd.github+json')
        rq.add_header('User-Agent', 'ps5-excel-publish')
        try:
            with urllib.request.urlopen(rq, timeout=300) as r:
                return r.status, json.loads(r.read().decode('utf-8'))
        except urllib.error.HTTPError as e:
            try:
                return e.code, json.loads(e.read().decode('utf-8'))
            except Exception:
                return e.code, {}

    st, js = req('GET')
    sha = js.get('sha') if st == 200 else None
    b64 = base64.b64encode(open(XLSX, 'rb').read()).decode('utf-8')
    body = json.dumps({
        'message': 'publish excel ' + time.strftime('%Y-%m-%d %H:%M'),
        'content': b64,
        'sha': sha,
        'branch': 'main',
    }).encode('utf-8')
    st2, js2 = req('PUT', body)
    if st2 in (200, 201):
        print('[publish] excel uploaded - the platform download button now '
              'serves this file')
        return True
    print(f'[publish] upload failed: HTTP {st2} - {js2.get("message")}')
    return False


data = None


def snapshot_from_workbook():
    """Record what the golden workbook has (for diff next run)."""
    snap = {}
    wb = openpyxl.load_workbook(XLSX, data_only=False)
    sidmap = sidmap_from_wb(wb)
    for sh_name, start, idc, edmap, cloud_sh, notes_col, color_col in SHEETS_FOR_PUSH:
        if sh_name not in wb.sheetnames:
            continue
        ws = wb[sh_name]
        for r in range(start, ws.max_row + 1):
            key = ws.cell(r, idc).value
            if not key:
                continue
            key = str(key).strip()
            idv = (sidmap.get(key, key)
                   if sh_name in SIDMAP_SHEETS else key)
            snap.setdefault(sh_name, {})[idv] = {}
            for col, pcol_name in edmap.items():
                snap[sh_name][idv][pcol_name] = norm(ws.cell(r, col).value)
            cc = ws.cell(r, color_col).fill
            cur = ''
            try:
                if cc and cc.fill_type == 'solid' and cc.start_color and \
                        cc.start_color.rgb and str(cc.start_color.rgb) != '00000000':
                    raw = str(cc.start_color.rgb)
                    cur = ('#FF' + raw[-6:].upper()) if len(raw) == 8 else ('#FF' + raw.upper())
            except Exception:
                cur = ''
            snap[sh_name][idv]['__COLOR__'] = cur
    wb.close()
    return snap


# -------------------------------------------------------------------- main --
def main():
    global data
    args = set(sys.argv[1:])
    token = ensure_token() if (('--sync' in args or '--push' in args
                                or '--publish' in args)
                               and '--nopush' not in args) else ''

    snapshot = read_snapshot()

    # 1) push local Excel edits first (so they win / merge online)
    if '--push' in args or '--sync' in args:
        changes = read_xlsx_edits(snapshot)
        if changes is None:
            print('[push] no workbook yet - nothing to push')
        else:
            n_cells = sum(len(v) for ids in changes['cells'].values()
                          for v in ids.values())
            n_notes = sum(len(v) for v in changes['notes'].values())
            n_cols = sum(len(ids) for ids in changes['colors'].values())
            if not (n_cells or n_notes or n_cols):
                print('[push] no new edits in the Excel - nothing to send')
            else:
                print(f'[push] sending {n_cells} cell(s), {n_notes} note(s), '
                      f'{n_cols} color(s)')
                push_state(changes, token)

    # 2) publish-only mode: upload the current xlsx to the repo, no rebuild
    if '--publish' in args and '--push' not in args and '--sync' not in args:
        publish_xlsx(token)
        return 0

    # 2b) per-page files mode: slice the golden workbook into 9 icons/files
    if '--pages' in args and '--push' not in args and '--sync' not in args:
        state = {}
        if '--local' not in args:
            state = get_state()
            if not state:
                print('[pull] cloud online state is empty/failed - using '
                      'platform data as-is')
        d9 = load_data_local(state)
        if d9 is None:
            return 1
        build_pages(d9)
        return 0

    # 2) pull live data + rebuild
    if '--push' in args and '--rebuild' not in args:
        return 0        # push-only mode
    state = get_state()
    if not state:
        print('[pull] cloud online state is empty/failed - using platform data '
              'as-is')
    data = load_data_local(state)
    if data is None:
        return 1
    SUBS, PUNCH, ITR, RFC, ITRT, PUNT, MILES, CABLES, OVR, NOTES, PCOL = data
    print(f'[data] SUBS={len(SUBS)} PUNCH={len(PUNCH)} ITR={len(ITR)} '
          f'RFC={len(RFC)} CABLES={len(CABLES)}')
    if not build_and_save():
        return 1
    write_snapshot(snapshot_from_workbook())
    print(f'[save] written: {os.path.basename(XLSX)}')

    if token and ('--publish' in args or '--sync' in args):
        publish_xlsx(token)

    if '--open' in sys.argv or '--sync' in args:
        try:
            import subprocess
            subprocess.Popen(['cmd', '/c', 'start', '', XLSX],
                             creationflags=0x08000000)
            print('[open] opened in Excel')
        except Exception:
            try:
                os.startfile(XLSX)
                print('[open] opened in Excel')
            except Exception:
                print('[open] could not auto-open - open the file yourself')
    print('DONE.')


if __name__ == '__main__':
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print('\ninterrupted.')
        sys.exit(130)