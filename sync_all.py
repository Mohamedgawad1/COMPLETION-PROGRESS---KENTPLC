"""sync_all.py - one-touch full sync for the PS5 platform.

Pipeline:
  1) rebuild_data.py            -> index.html from the newest DPR SUMMARY + ITR/Punch
  2) make_platform_excel.py     -> golden PS5 PLATFORM.xlsx from the platform data
  3) make_platform_excel.py --pages --local -> PAGES/*.xlsx (all 9 download pages)
  4) copy PAGES/* -> EXCEL/*  (this is what the download buttons serve)
  5) refresh EXCEL/7 - COMPLETE.xlsm (34 rows) while preserving the (+ ROW) macro
  6) git add/commit/push the changed files (EXCEL/*, PS5 PLATFORM.xlsx,
     index.html, rebuild_data.py, make_platform_excel.py)

Run from the repo folder:  python sync_all.py
"""
import os
import shutil
import subprocess
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.path.join(HERE, "EXCEL")
PAGES = os.path.join(HERE, "PAGES")
GOLDEN = os.path.join(HERE, "PS5 PLATFORM.xlsx")


def run(cmd):
    print("\n>>>", " ".join(cmd))
    r = subprocess.run(cmd, cwd=HERE)
    if r.returncode != 0:
        print("!! FAILED:", cmd, "exit", r.returncode)
        sys.exit(r.returncode)


def refresh_xlsm():
    import openpyxl
    new = os.path.join(PAGES, "7 - COMPLETE.xlsx")
    old = os.path.join(EXCEL, "7 - COMPLETE.xlsm")
    if not (os.path.exists(new) and os.path.exists(old)):
        print("  (xlsm refresh skipped - files missing)")
        return
    wb_new = openpyxl.load_workbook(new, data_only=True)
    rows = [list(r) for r in wb_new["7 \u00b7 COMPLETE"].iter_rows(min_row=2, values_only=True)]
    wb_new.close()
    wb = openpyxl.load_workbook(old, keep_vba=True, data_only=True)
    ws = wb["7 \u00b7 COMPLETE"]
    for i in range(ws.max_row, 1, -1):
        ws.delete_rows(i)
    for ri, r in enumerate(rows, start=2):
        for ci, v in enumerate(r, start=1):
            ws.cell(row=ri, column=ci, value=v)
    wb.save(old)
    wb.close()
    print(f"  [xlsm] 7 - COMPLETE.xlsm refreshed -> {len(rows)} rows, macro kept")


def copy_pages():
    os.makedirs(EXCEL, exist_ok=True)
    for f in sorted(os.listdir(PAGES)):
        if f.lower().endswith(".xlsx"):
            shutil.copy2(os.path.join(PAGES, f), os.path.join(EXCEL, f))
    print("  [files] PAGES/* -> EXCEL/*")


def main():
    run([sys.executable, "rebuild_data.py"])
    run([sys.executable, "make_platform_excel.py"])
    run([sys.executable, "make_platform_excel.py", "--pages", "--local"])
    copy_pages()
    refresh_xlsm()

    print("\n>>> git pull (clean tree) / commit only-if-changed / push")
    subprocess.run(["git", "pull", "--rebase", "origin", "main"],
                   cwd=HERE, check=True)
    subprocess.run(["git", "add", "--", "EXCEL", GOLDEN, "index.html",
                    "rebuild_data.py", "make_platform_excel.py"],
                   cwd=HERE, check=True)
    staged = subprocess.run(["git", "diff", "--cached", "--quiet"], cwd=HERE).returncode
    if staged != 0:
        subprocess.run(["git", "commit", "-m", "Full sync: platform + all Excel downloads"],
                       cwd=HERE, check=True)
        subprocess.run(["git", "push", "origin", "main"], cwd=HERE, check=True)
    else:
        print("  nothing staged to commit - already in sync")
    print("\nDONE - platform online + all downloads synced.")


if __name__ == "__main__":
    main()