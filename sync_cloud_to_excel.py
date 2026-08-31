# -*- coding: utf-8 -*-
"""
PS5 - Sync platform edits (GitHub Pages) -> PS-5 COMPLETIONS DPR SUMMERY Excel.

Pulls https://mohamedgawad1.github.io/PS5-COMPLETION-PLATFORM/platform_state.json
and applies every platform edit to the DPR SUMMERY workbook with EXACT mapping:

  PUNCH LIST -> sheet 'DETAILED PUNCH LIST'  (id col 'Punchlist ID')
  ITR LIST   -> sheet 'DETAILED ITR LIST'    (id col 'Task ID')
  RFC PROGRESS -> sheet 'RFC PROGRESS'       (id = text before ' - ' in col A,
                                              data starts row 4)

Column map is fixed below (platform name -> exact Excel header). Nothing is
written into formula cells (BALANCE / % / ITRs / CLOSED totals are formulas and
recalculate automatically). Notes go to a 'PLATFORM NOTES' column that is
created at the far right if missing. Row colors become the fill of the ID cell.

Usage:
  python sync_cloud_to_excel.py            # one shot
  python sync_cloud_to_excel.py --watch    # repeat every 5 minutes
  python sync_cloud_to_excel.py --force    # re-apply even if state unchanged
"""
import hashlib
import json
import os
import re
import shutil
import sys
import time
import urllib.request
import zipfile

import openpyxl
from openpyxl.styles import Font, PatternFill

HOME = os.path.join(os.path.expanduser('~'), 'Downloads')
DL_SUB = os.path.join(HOME, 'PS5 - CPP AGI Completion Progress Dashboard_files')
HERE = os.path.dirname(os.path.abspath(__file__))

STATE_URL = ('https://raw.githubusercontent.com/Mohamedgawad1/'
             'PS5-COMPLETION-PLATFORM/main/platform_state.json')
SEEN_FILE = os.path.join(HERE, '_platform_sync_seen.json')
REPORT_FILE = os.path.join(HERE, '_platform_sync_report.txt')

# ---------------------------------------------------------------- mapping --
# platform column  -> exact Excel header text
PUNCH_MAP = {
    'TAG': 'Asset (Name/Tag)',
    'CAT': 'CAT',
    'DISC': 'Discipline (Name)',
    'DESCRIPTION': 'Description',
    'STATUS': 'Status',
    'CLOSING DATE': 'Workflow - Closing Date',
}
ITR_MAP = {
    'TAG': 'Asset - Tag',
    'DISC': 'Discipline',
    'TASK TYPE': 'Task Type (Name)',
    'ASSET DESCRIPTION': 'Asset - Description',
    'STATE': 'Task State',
    'CLOSING DATE': 'Closing Date',
}
# RFC PROGRESS sheet: positional (1-based) columns, verified against the file
RFC_ID_COL = 1          # 'PS5-01-01 - description'
RFC_DATA_ROW = 4        # first data row
# platform column -> fallback column (only used when header scan fails)
RFC_MAP = {
    'Priority': 2,
    'RFC BHMPS': 3,
    'RFC EIT': 4,
    'Baseline': 5,
    'Recovery': 8,
    'SIGNED': 9,
    'Milestone': 10,
    'EIT': 56,
    'EACOP': 55,
    'REMARK EACOP': 55,
    'REMARK CPP-EIT': 56,
    'REMARK CPP-1': 57,
    'STATUS': None,      # resolved dynamically -> 'WALKDOWN STATUS' column
}
# header hints used to locate the REAL column in the current DPR workbook
# (headers drift daily, so fixed indexes are unreliable)
RFC_HINTS = {
    'Priority': ('PRIORITY',),
    'RFC BHMPS': ('BHMPS', 'RFSU'),
    'RFC EIT': ('DESCIPILINE',),
    'Baseline': ('CPP 1',),
    'Recovery': ('RECOVERY',),
    'SIGNED': ('RFC SIGNED',),
    'Milestone': ('MILESTONE',),
    'EIT': ('CPP-EIT',),
    'EACOP': ('CPY', 'EACOP'),
    'REMARK EACOP': ('BLOCKING POINTS REMARKS',),
    'REMARK CPP-EIT': ('CPP-EIT',),
    'REMARK CPP-1': ('CPP-1',),
}
# discipline letter -> (TOTAL col, CLOSED col) for THIS workbook layout
RFC_DISC_COLS = {'B': (12, 13), 'E': (16, 17), 'H': (20, 21), 'I': (24, 25),
                 'M': (28, 29), 'P': (32, 33), 'S': (36, 37), 'T': (40, 41)}
# platform columns that are derived in Excel (formulas) - never written
RFC_DERIVED = {'ITRs', 'CLOSED', 'BALANCE'}


def rfc_columns(ws):
    """Locate platform column -> Excel column by scanning header rows 1-3."""
    out = {}
    for r in range(1, 4):
        for c in range(1, min(ws.max_column, 90) + 1):
            t = re.sub(r'\s+', ' ', n(ws.cell(r, c).value)).upper()
            if not t:
                continue
            for key, hints in RFC_HINTS.items():
                if key in out:
                    continue
                for h in hints:
                    if h in t:
                        out[key] = c
                        break
    return out

NOTE_HEADER = 'PLATFORM NOTES'
TARGET_SHEET = {
    'PUNCH LIST': ('DETAILED PUNCH LIST', 'Punchlist ID', PUNCH_MAP),
    'ITR LIST': ('DETAILED ITR LIST', 'Task ID', ITR_MAP),
}


def n(v):
    return '' if v is None else str(v).strip()


def fetch_state(path=None):
    if path:
        try:
            with open(path, encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f'[cloud] state file read failed: {e}')
            return None
    try:
        req = urllib.request.Request(STATE_URL, headers={'User-Agent': 'ps5-sync'})
        with urllib.request.urlopen(req, timeout=30) as r:
            return json.loads(r.read().decode('utf-8'))
    except Exception as e:
        print(f'[cloud] download failed: {e}')
        return None


def load_seen():
    try:
        with open(SEEN_FILE, encoding='utf-8') as f:
            return json.load(f).get('hash')
    except Exception:
        return None


def save_seen(h):
    with open(SEEN_FILE, 'w', encoding='utf-8') as f:
        json.dump({'hash': h, 'time': time.strftime('%Y-%m-%d %H:%M:%S')}, f)


MONTH_NAMES = {'JANUARY': '01', 'FEBRUARY': '02', 'MARCH': '03', 'APRIL': '04',
               'MAY': '05', 'JUNE': '06', 'JULY': '07', 'AUGUST': '08',
               'SEPTEMBER': '09', 'OCTOBER': '10', 'NOVEMBER': '11',
               'DECEMBER': '12'}
DATE_RE = re.compile(r'PS-5\s*COMPLETIONS\s*DPR\s*SUMMERY\s*-\s*'
                     r'(\d{1,2})-(\d{1,2})-(\d{2,4})\.xlsx$', re.I)
DATE_RE_MONTH = re.compile(r'PS-5\s*COMPLETIONS\s*DPR\s*SUMMERY\s*-\s*'
                           r'(\d{1,2})-([A-Z]+)-(\d{2,4})\.xlsx$', re.I)


def _fname_date(f):
    """(year, month, day) from '-DD-MM-YY' or '-DD-MONTHNAME-YY' suffix,
    or None when the name carries no date."""
    m = DATE_RE.search(f)
    if m:
        dd, mm, yy = m.group(1), m.group(2), m.group(3)
        yy = yy if len(yy) == 4 else '20' + yy
        return (yy, mm.zfill(2), dd.zfill(2))
    m = DATE_RE_MONTH.search(f)
    if m and m.group(2).upper() in MONTH_NAMES:
        dd, mm, yy = m.group(1), MONTH_NAMES[m.group(2).upper()], m.group(3)
        yy = yy if len(yy) == 4 else '20' + yy
        return (yy, mm, dd.zfill(2))
    return None


def find_target():
    """Newest DPR SUMMERY by the DATE SUFFIX in its name ('-DD-MM-YY' or
    '-DD-MONTHNAME-YY'); an undated file falls back to its mtime."""
    best = None
    seen = set()
    for folder in (HOME, DL_SUB):
        if not os.path.isdir(folder):
            continue
        for f in os.listdir(folder):
            b = f.upper()
            if (f.startswith('~$') or not f.lower().endswith('.xlsx')
                    or 'BACKUP' in b):
                continue
            p = os.path.join(folder, f)
            key = os.path.realpath(p)
            if key in seen:
                continue
            seen.add(key)
            dk = _fname_date(f)
            cand = (dk if dk else ('0000', '00', '00'),
                    os.path.getmtime(p), p)
            if best is None or cand[:2] > best[:2]:
                best = cand
    return best[2] if best else None


def header_index(ws, title):
    """Exact (case/space-insensitive) header lookup in row 1."""
    want = re.sub(r'\s+', ' ', n(title)).upper()
    for c in range(1, ws.max_column + 1):
        if re.sub(r'\s+', ' ', n(ws.cell(1, c).value)).upper() == want:
            return c
    return None


def ensure_note_col(ws):
    c = header_index(ws, NOTE_HEADER)
    if c:
        return c
    c = ws.max_column + 1
    ws.cell(1, c).value = NOTE_HEADER
    return c


def is_formula(cell):
    v = cell.value
    if isinstance(v, str) and v.startswith('='):
        return True
    return type(v).__name__ == 'ArrayFormula'


def build_rfc_ids(ws):
    ids = {}
    for r in range(RFC_DATA_ROW, ws.max_row + 1):
        raw = n(ws.cell(r, RFC_ID_COL).value)
        if not raw:
            continue
        sid = raw.split(' - ')[0].strip().upper()
        if sid and sid not in ids:
            ids[sid] = r
    return ids


def apply_fill(ws, row, col, color):
    cell = ws.cell(row, col)
    if color:
        argb = color if len(color) == 8 else 'FF' + color.lstrip('#')
        cell.fill = PatternFill(start_color=argb, end_color=argb,
                                fill_type='solid')


def index_ids(ws, idc, start=2):
    """Map ID -> row. Stops after a long run of empty rows (formatted-but-empty
    sheets like DETAILED ITR LIST report max_row = 1048576)."""
    rows_by_id = {}
    empty = 0
    for r in range(start, ws.max_row + 1):
        rid = n(ws.cell(r, idc).value)
        if rid:
            rows_by_id[rid.upper()] = r
            empty = 0
        else:
            empty += 1
            if empty > 300:
                break
    return rows_by_id


def main():
    args = sys.argv[1:]
    force = '--force' in args
    watch = '--watch' in args
    state_file = None
    target = None
    if '--state' in args:
        state_file = args[args.index('--state') + 1]
    if '--target' in args:
        target = args[args.index('--target') + 1]
    while True:
        try:
            run_once(force, state_file, target)
        except Exception as e:
            print('[sync] error:', e)
        if not watch:
            return 0
        time.sleep(300)


def run_once(force, state_file=None, target=None):
    global SEEN_FILE
    st = fetch_state(state_file)
    if st is None:
        return
    if state_file:
        force = True
    blob = json.dumps(st, sort_keys=True, ensure_ascii=False)
    h = hashlib.sha256(blob.encode()).hexdigest()
    if h == load_seen() and not force:
        print('[sync] cloud state unchanged - nothing to do')
        return

    src = target or find_target()
    if not src:
        print('[sync] no DPR SUMMERY workbook found!')
        return
    print(f'[sync] target : {os.path.basename(src)}')

    cells = st.get('cells', {}) or {}
    notes = st.get('notes', {}) or {}
    colors = st.get('colors', {}) or {}

    backup = os.path.join(os.path.dirname(src),
                          os.path.splitext(os.path.basename(src))[0]
                          + ' -BACKUP.xlsx')
    shutil.copy2(src, backup)

    try:
        wb = openpyxl.load_workbook(src)
    except Exception as e:
        print(f'[sync] cannot open "{os.path.basename(src)}" ({e}) - '
              f'close Excel / check the file. NOTHING was changed.')
        return
    rep = {'written': 0, 'notes': 0, 'colors': 0, 'skipped_formula': [],
           'skipped_derived': [], 'no_row': [], 'no_col': []}

    # ---- detailed sheets -------------------------------------------------
    for psheet, (tgt, idh, cmap) in TARGET_SHEET.items():
        has_edits = (cells.get(psheet) or notes.get(psheet)
                     or colors.get(psheet))
        if tgt not in wb.sheetnames or not has_edits:
            continue
        ws = wb[tgt]
        idc = header_index(ws, idh)
        if not idc:
            rep['no_col'].append(f'{tgt}: ID column {idh!r} missing')
            continue
        rows_by_id = index_ids(ws, idc)
        colmap = {}
        for pcol, etitle in cmap.items():
            ec = header_index(ws, etitle)
            if ec:
                colmap[pcol] = ec
            else:
                rep['no_col'].append(f'{tgt}: {etitle!r} missing')
        note_c = ensure_note_col(ws) if notes.get(psheet) else None

        for rid, ed in (cells.get(psheet, {}) or {}).items():
            r = rows_by_id.get(n(rid).upper())
            if not r:
                rep['no_row'].append(f'{psheet}/{rid}')
                continue
            for pcol, val in ed.items():
                ec = colmap.get(pcol)
                if not ec:
                    rep['no_col'].append(f'{psheet}: {pcol} unmapped')
                    continue
                cell = ws.cell(r, ec)
                if is_formula(cell):
                    rep['skipped_formula'].append(f'{tgt}!{cell.coordinate}')
                    continue
                cell.value = val
                cell.font = Font(size=9)
                rep['written'] += 1
        for rid, val in (notes.get(psheet, {}) or {}).items():
            r = rows_by_id.get(n(rid).upper())
            if not r:
                rep['no_row'].append(f'{psheet}-note/{rid}')
                continue
            ws.cell(r, note_c).value = val
            ws.cell(r, note_c).font = Font(size=9)
            rep['notes'] += 1
        for rid, col in (colors.get(psheet, {}) or {}).items():
            r = rows_by_id.get(n(rid).upper())
            if r:
                apply_fill(ws, r, 1, col)
                rep['colors'] += 1

    # ---- RFC PROGRESS ----------------------------------------------------
    sh = 'RFC PROGRESS'
    adds = st.get('adds', {}) or {}
    deladds = st.get('deladds', {}) or {}
    if sh in wb.sheetnames and (cells.get(sh) or notes.get(sh)
                                or colors.get(sh) or adds or deladds):
        ws = wb[sh]
        rfc_cols = rfc_columns(ws)

        def rfc_col(pcol):
            if pcol in rfc_cols:
                return rfc_cols[pcol]
            return RFC_MAP.get(pcol)

        # platform sid = text before ' - ' in column A
        ids = {k.split(' - ')[0].strip(): v
               for k, v in index_ids(ws, RFC_ID_COL,
                                     start=RFC_DATA_ROW).items()}
        walk_holder = [None]

        def get_walk():
            if walk_holder[0] is None:
                c = None
                for cc in range(56, ws.max_column + 1):
                    if re.sub(r'\s+', ' ', n(ws.cell(3, cc).value)).upper() \
                            == 'WALKDOWN STATUS':
                        c = cc
                        break
                if c is None:
                    c = max(ws.max_column, 55) + 1
                    ws.cell(3, c).value = 'WALKDOWN STATUS'
                walk_holder[0] = c
            return walk_holder[0]

        def put(r, cidx, val):
            cell = ws.cell(r, cidx)
            if is_formula(cell):
                rep['skipped_formula'].append(f'{sh}!{cell.coordinate}')
                return
            cell.value = val
            cell.font = Font(size=9)
            rep['written'] += 1

        for rid, ed in (cells.get(sh, {}) or {}).items():
            r = ids.get(n(rid).upper())
            if not r:
                rep['no_row'].append(f'{sh}/{rid}')
                continue
            for pcol, val in ed.items():
                mkey = re.match(r'^([A-T]) (TOTAL|CLOSED)$', pcol)
                if mkey and mkey.group(1) in RFC_DISC_COLS:
                    tc, cc = RFC_DISC_COLS[mkey.group(1)]
                    put(r, tc if mkey.group(2) == 'TOTAL' else cc, val)
                elif pcol in RFC_DERIVED:
                    rep['skipped_derived'].append(f'{sh}:{pcol} ({rid})')
                elif pcol in RFC_MAP or pcol in rfc_cols:
                    cidx = rfc_col(pcol)
                    if cidx is None and pcol == 'STATUS':
                        cidx = get_walk()
                    if cidx:
                        put(r, cidx, val)
                    else:
                        rep['no_col'].append(f'{sh}: {pcol} unmapped')
                else:
                    rep['no_col'].append(f'{sh}: {pcol} unmapped')
        for rid, val in (notes.get(sh, {}) or {}).items():
            r = ids.get(n(rid).upper())
            if not r:
                rep['no_row'].append(f'{sh}-note/{rid}')
                continue
            nc = ensure_note_col(ws)
            ws.cell(r, nc).value = val
            ws.cell(r, nc).font = Font(size=9)
            rep['notes'] += 1
        for rid, col in (colors.get(sh, {}) or {}).items():
            r = ids.get(n(rid).upper())
            if r:
                apply_fill(ws, r, 1, col)
                rep['colors'] += 1

        # ---- online ADDED rows ('adds') & DELETE tombstones ('deladds') --
        if adds or deladds:
            base_sids = set(_load_rfc_from_platform())

            def find_row(sid):
                for r in range(RFC_DATA_ROW, ws.max_row + 1):
                    if n(ws.cell(r, RFC_ID_COL).value
                         ).split(' - ')[0].strip().upper() == sid.upper():
                        return r
                return None

            def last_data_row():
                last = RFC_DATA_ROW
                empty = 0
                for rr in range(RFC_DATA_ROW, ws.max_row + 1):
                    if n(ws.cell(rr, RFC_ID_COL).value):
                        last = rr
                        empty = 0
                    else:
                        empty += 1
                        if empty > 20:
                            break
                return last or RFC_DATA_ROW

            def put_rfc(row, pcol, val):
                if val is None or val == '':
                    return
                cidx = rfc_col(pcol)
                if not cidx:
                    return
                cell = ws.cell(row, cidx)
                if is_formula(cell):
                    rep['skipped_formula'].append(f'{sh}!{cell.coordinate}')
                    return
                cell.value = val
                cell.font = Font(size=9)
                rep['written'] += 1

            for sid in deladds:
                sid = n(sid)
                if not sid or sid in base_sids:
                    continue
                r = find_row(sid)
                if r:
                    ws.delete_rows(r)
                    rep['written'] += 1
                    print(f'[sync] deladds: removed {sid} from {sh}')

            for sid, a in (adds or {}).items():
                sid = n(sid)
                nm = (a.get('name') or '').strip()
                if not sid or sid in deladds or sid in base_sids:
                    continue
                if find_row(sid):
                    continue
                row = last_data_row() + 1
                ws.cell(row, RFC_ID_COL).value = (sid + ' - ' + nm).strip()
                ws.cell(row, RFC_ID_COL).font = Font(size=9)
                rep['written'] += 1
                put_rfc(row, 'Priority', a.get('prio'))
                put_rfc(row, 'RFC BHMPS', a.get('bhm'))
                put_rfc(row, 'RFC EIT', a.get('eit'))
                put_rfc(row, 'Baseline', a.get('base'))
                put_rfc(row, 'Recovery', a.get('rec'))
                put_rfc(row, 'SIGNED', a.get('signed'))
                put_rfc(row, 'Milestone', a.get('mile'))
                put_rfc(row, 'REMARK EACOP', a.get('re1'))
                put_rfc(row, 'REMARK CPP-EIT', a.get('re2'))
                put_rfc(row, 'REMARK CPP-1', a.get('re3'))
                print(f'[sync] adds: appended {sid} to {sh}')

    # ---- nothing changed? do not touch the workbook at all ---------------
    if not (rep['written'] or rep['notes'] or rep['colors']):
        wb.close()
        save_seen(h) if not state_file else None
        print('[sync] no effective changes - workbook untouched')
        return

    # ---- atomic save: tmp file -> validate -> replace --------------------
    tmpf = src + '.saving.tmp'
    ok = False
    for attempt in range(15):
        try:
            wb.save(tmpf)
            ok = True
            break
        except PermissionError:
            print(f'[wait] close "{os.path.basename(src)}" in Excel '
                  f'({attempt + 1}/15)...')
            time.sleep(2)
    wb.close()
    if not ok:
        if os.path.exists(tmpf):
            os.remove(tmpf)
        print('[sync] file locked - nothing changed, try again later')
        return
    try:
        _z = zipfile.ZipFile(tmpf)
        _bad = _z.testzip()
        _n = len(_z.namelist())
        _z.close()
        if _bad is not None or _n < 10:
            raise RuntimeError(f'invalid archive ({_n} entries)')
    except Exception as e:
        if os.path.exists(tmpf):
            os.remove(tmpf)
        print(f'[sync] post-save validation FAILED ({e}) - '
              f'original file untouched')
        return
    os.replace(tmpf, src)

    save_seen(h) if not state_file else None
    lines = [
        f"sync {time.strftime('%Y-%m-%d %H:%M:%S')} -> {os.path.basename(src)}",
        f"written={rep['written']} notes={rep['notes']} "
        f"colors={rep['colors']}"]
    if rep['skipped_derived']:
        lines.append('derived(auto in Excel, not written): '
                     + ', '.join(rep['skipped_derived'][:20]))
    if rep['skipped_formula']:
        lines.append('formula cells protected: '
                     + ', '.join(rep['skipped_formula'][:20]))
    if rep['no_row']:
        lines.append('rows NOT found: ' + ', '.join(rep['no_row'][:30]))
    if rep['no_col']:
        lines.append('columns missing/unmapped: '
                     + ', '.join(sorted(set(rep['no_col']))[:20]))
    text = '\n'.join(lines)
    print(text)
    with open(REPORT_FILE, 'w', encoding='utf-8') as f:
        f.write(text + '\n')


if __name__ == '__main__':
    raise SystemExit(main())
