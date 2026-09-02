"""
GitHub Actions sync: platform_state.json -> Excel
Runs every 5 minutes via GitHub Actions workflow.
"""
import json
import os
import sys
import glob
import requests
import openpyxl
from datetime import datetime

# Config
REPO_OWNER = os.environ.get('REPO_OWNER', 'Mohamedgawad1')
REPO_NAME = os.environ.get('REPO_NAME', 'PS5-COMPLETION-PLATFORM')
TOKEN = os.environ.get('GITHUB_TOKEN', '')
EXCEL_PATTERN = 'PS-5 COMPLETIONS DPR SUMMERY*.xlsx'

def log(msg):
    print(f'[{datetime.now().strftime("%H:%M:%S")}] {msg}')

def fetch_state():
    """Fetch platform_state.json from GitHub API"""
    url = f'https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/platform_state.json'
    headers = {'Authorization': f'token {TOKEN}', 'Accept': 'application/vnd.github.raw+json'}
    
    r = requests.get(url, headers=headers, timeout=30)
    if r.status_code == 200:
        log(f'Fetched state: {len(r.content)} bytes')
        return r.json()
    else:
        log(f'Failed to fetch state: {r.status_code}')
        return None

def find_excel():
    """Find latest DPR SUMMERY file"""
    files = sorted(glob.glob(EXCEL_PATTERN), key=os.path.getmtime, reverse=True)
    return files[0] if files else None

def apply_edits(state, excel_path):
    """Apply platform edits to Excel"""
    log(f'Opening Excel: {excel_path}')
    wb = openpyxl.load_workbook(excel_path)
    
    edits = state.get('edits', {})
    notes = state.get('notes', {})
    colors = state.get('colors', {})
    prc = state.get('prc', {})
    
    applied = 0
    
    # Apply edits to sheets
    for sheet_name, cells in edits.items():
        if sheet_name not in wb.sheetnames:
            log(f'Sheet not found: {sheet_name}')
            continue
        ws = wb[sheet_name]
        for cell_ref, value in cells.items():
            try:
                ws[cell_ref] = value
                applied += 1
            except Exception as e:
                log(f'Error setting {sheet_name}!{cell_ref}: {e}')
    
    # Apply notes
    for sheet_name, cells in notes.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for cell_ref, value in cells.items():
            try:
                ws[cell_ref] = value
                applied += 1
            except:
                pass
    
    # Apply colors
    from openpyxl.styles import PatternFill
    for sheet_name, cells in colors.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for cell_ref, color in cells.items():
            try:
                fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                ws[cell_ref].fill = fill
                applied += 1
            except:
                pass
    
    # Apply prc
    for sheet_name, cells in prc.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for cell_ref, value in cells.items():
            try:
                ws[cell_ref] = value
                applied += 1
            except:
                pass
    
    if applied > 0:
        wb.save(excel_path)
        log(f'Saved: {applied} edits applied')
    else:
        log('No edits to apply')
    
    return applied

def main():
    if not TOKEN:
        log('ERROR: GITHUB_TOKEN not set')
        sys.exit(1)
    
    # Fetch state
    state = fetch_state()
    if state is None:
        log('No state found - exiting')
        sys.exit(0)
    
    # Find Excel
    excel = find_excel()
    if not excel:
        log('ERROR: No DPR SUMMERY file found')
        sys.exit(1)
    
    # Apply edits
    applied = apply_edits(state, excel)
    log(f'Done: {applied} edits synced')

if __name__ == '__main__':
    main()
