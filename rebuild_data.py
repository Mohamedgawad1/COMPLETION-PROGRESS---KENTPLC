"""
Rebuild ALL platform data from XLSX source files.
PUNCH, ITR, RFC, ITRT, PUNT, SUBS, CABLES — all regenerated.
RFC workflow fields from DPR SUMMARY.
"""
import json, re, sys, os
from collections import Counter, defaultdict
import openpyxl

print("="*60)
print("STEP 1: Read XLSX files")
print("="*60)

# --- ITR from ovTasks_TestsPlanned_1369.xlsx ---
print("\nReading ovTasks_TestsPlanned_1369.xlsx...")
wb_itr = openpyxl.load_workbook('C:/Users/mylap/Downloads/PS5 - CPP AGI Completion Progress Dashboard_files/ovTasks_TestsPlanned_1369.xlsx', read_only=True, data_only=True)
ws_itr = wb_itr['Exported from SC']

itr_raw = []
disc_totals = defaultdict(lambda: {'total': 0, 'closed': 0})
sub_totals = defaultdict(lambda: {'total': 0, 'closed': 0})
sub_plant = {}
sub_name_map = {}
plant_totals = defaultdict(lambda: {'total': 0, 'closed': 0})

for row in ws_itr.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    tid = str(row[0] or '')
    tag = str(row[1] or '')
    disc = str(row[8] or '')
    ttype = str(row[9] or '')
    desc = str(row[11] or '')
    state = str(row[28] or '')
    closedate = str(row[22] or '') if row[22] else ''
    company = str(row[13] or '')
    sub_full = str(row[21] or '')
    sub_id = str(row[23] or '')
    plant = str(row[12] or '')
    
    is_done = state.strip().lower().startswith('comp') or state.strip().lower() == 'closed'
    
    itr_raw.append([tid, tag, disc, ttype, desc, state, closedate, company, sub_full, sub_id])
    
    if sub_id:
        sub_totals[sub_id]['total'] += 1
        if is_done: sub_totals[sub_id]['closed'] += 1
        sub_plant[sub_id] = plant
        sub_name_map[sub_id] = sub_full
    
    disc_totals[disc]['total'] += 1
    if is_done: disc_totals[disc]['closed'] += 1
    
    if plant:
        plant_totals[plant]['total'] += 1
        if is_done: plant_totals[plant]['closed'] += 1

wb_itr.close()
print(f"  ITR tasks: {len(itr_raw)}")
print(f"  Unique subsystems: {len(sub_totals)}")

# Build asset tag → subsystem ID mapping from ITR
tag_to_sub = {}
for r in itr_raw:
    tag = r[1].strip()  # Asset Tag
    sid = r[9].strip()  # Subsystem ID
    if tag and sid:
        tag_to_sub[tag] = sid
print(f"  Asset tag → subsystem map: {len(tag_to_sub)} entries")

# --- Punch from ovPunchlist_1399.xlsx ---
print("\nReading ovPunchlist_1399.xlsx...")
wb_punch = openpyxl.load_workbook('C:/Users/mylap/Downloads/PS5 - CPP AGI Completion Progress Dashboard_files/ovPunchlist_1399.xlsx', data_only=True)
ws_punch = wb_punch.active

punch_raw = []
punch_disc_cat = defaultdict(lambda: defaultdict(lambda: {'total': 0, 'closed': 0}))
punch_statuses = Counter()

for row in ws_punch.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    pid = str(row[0] or '')
    tag = str(row[1] or '')
    cat = str(row[3] or '')
    disc = str(row[4] or '')
    desc = str(row[5] or '')
    sub_full = str(row[7] or '')
    closedate = str(row[8] or '') if row[8] else ''
    status = str(row[9] or '')
    action = str(row[11] or '') if row[11] else ''
    sub_id = sub_full.split(' - ')[0].strip() if ' - ' in sub_full else sub_full.strip()
    
    is_closed = status.strip().lower() in ('closed', 'completed')
    
    punch_raw.append([pid, tag, cat, disc, desc, status, action, 0, sub_full, sub_id])
    
    punch_statuses[status] += 1
    punch_disc_cat[disc][cat]['total'] += 1
    if is_closed:
        punch_disc_cat[disc][cat]['closed'] += 1

wb_punch.close()
print(f"  Punch items: {len(punch_raw)}")

print("\n" + "="*60)
print("STEP 2: Read existing HTML data (MILES, OVR)")
print("="*60)

with open('C:/Users/mylap/OneDrive/Desktop/PS5-COMPLETION-PLATFORM/index.html', 'r', encoding='utf-8') as f:
    html = f.read()

lines = html.split('\n')

# Find data line dynamically (search for 'const SUBS=')
data_line = ''
data_line_idx = -1
for i, ln in enumerate(lines):
    if 'const SUBS=' in ln:
        data_line = ln
        data_line_idx = i
        break

# Find OVR/NOTES/PCOL line (search for 'const NOTES=' or 'OVR=' on the line AFTER data)
ovr_line = ''
for i in range(data_line_idx + 1, min(data_line_idx + 5, len(lines))):
    if 'OVR=' in lines[i] or 'NOTES=' in lines[i]:
        ovr_line = lines[i]
        break

def extract_array(line, name):
    idx = line.find(f'{name}=[')
    if idx < 0: return []
    start = idx + len(name) + 1
    depth = 0
    for i in range(start, len(line)):
        if line[i] == '[': depth += 1
        elif line[i] == ']':
            depth -= 1
            if depth == 0: return json.loads(line[start:i+1])
    return []

def extract_obj(line, name):
    idx = line.find(f'{name}=')
    if idx < 0: return {}
    start = idx + len(name) + 1
    depth = 0
    for i in range(start, len(line)):
        if line[i] == '{': depth += 1
        elif line[i] == '}':
            depth -= 1
            if depth == 0: return json.loads(line[start:i+1])
    return {}

print(f"  Data line index: {data_line_idx + 1}")
print(f"  Data line size: {len(data_line)} chars")

MILES = extract_array(data_line, 'MILES')
OVR = extract_obj(ovr_line, 'OVR')
NOTES_raw = extract_obj(ovr_line, 'NOTES')
PCOL_raw = extract_obj(ovr_line, 'PCOL')

print(f"  MILES: {len(MILES)}")
print(f"  OVR sheets: {list(OVR.keys())}")

print("\n" + "="*60)
print("STEP 3: Read RFC PROGRESS from DPR SUMMARY")
print("="*60)

import glob as globmod
dpr_dir = 'C:/Users/mylap/Downloads/PS5 - CPP AGI Completion Progress Dashboard_files'
dpr_files = sorted(globmod.glob(os.path.join(dpr_dir, 'PS-5 COMPLETIONS DPR SUMMERY -*.xlsx')),
                   key=os.path.getmtime, reverse=True)

dpr_data = {}  # subsystem_id -> {priority, bhm, eit, base, rec, signed, mile, b1, b2, b3, re1, re2, re3, wd}

if dpr_files:
    dpr_name = os.path.basename(dpr_files[0])
    print(f"  DPR file: {dpr_name}")
    wb_dpr = openpyxl.load_workbook(dpr_files[0], read_only=True, data_only=True)
    ws_dpr = wb_dpr['RFC PROGRESS']
    
    for row in ws_dpr.iter_rows(min_row=4, values_only=True):
        vals = list(row[:81])
        sub = str(vals[0] or '').strip()
        if not sub or 'GRAND TOTAL' in sub.upper():
            continue
        
        # Extract subsystem ID from "PS5-01-01 - Fire Water Jockey Pump A/B"
        sid = sub.split(' - ')[0].strip() if ' - ' in sub else sub.strip()
        
        def fmt_date(v):
            if v is None: return ''
            s = str(v).strip()
            if not s or s == 'None': return ''
            # Format datetime
            if hasattr(v, 'strftime'):
                return v.strftime('%Y-%m-%d')
            return s
        
        def fmt_num(v):
            if v is None: return 0
            if isinstance(v, (int, float)): return int(v)
            s = str(v).strip()
            if not s or s == 'None': return 0
            try: return int(float(s))
            except: return 0
        
        def fmt_str(v):
            if v is None: return ''
            s = str(v).strip()
            if s == 'None': return ''
            return s
        
        dpr_data[sid] = {
            'priority': fmt_str(vals[1]),
            'bhm': fmt_date(vals[2]),
            'eit': fmt_date(vals[3]),
            'base': fmt_date(vals[4]),
            'rec': fmt_date(vals[5]),
            'signed': fmt_date(vals[6]),
            'mile': fmt_str(vals[7]),
            'tot_pct': vals[8],  # keep raw for now
            'b_total': fmt_num(vals[9]),
            'b_closed': fmt_num(vals[10]),
            'b_bal': fmt_num(vals[11]),
            'e_total': fmt_num(vals[13]),
            'e_closed': fmt_num(vals[14]),
            'e_bal': fmt_num(vals[15]),
            'h_total': fmt_num(vals[17]),
            'h_closed': fmt_num(vals[18]),
            'h_bal': fmt_num(vals[19]),
            'i_total': fmt_num(vals[21]),
            'i_closed': fmt_num(vals[22]),
            'i_bal': fmt_num(vals[23]),
            'm_total': fmt_num(vals[25]),
            'm_closed': fmt_num(vals[26]),
            'm_bal': fmt_num(vals[27]),
            'p_total': fmt_num(vals[29]),
            'p_closed': fmt_num(vals[30]),
            'p_bal': fmt_num(vals[31]),
            's_total': fmt_num(vals[33]),
            's_closed': fmt_num(vals[34]),
            's_bal': fmt_num(vals[35]),
            't_total': fmt_num(vals[37]),
            't_closed': fmt_num(vals[38]),
            't_bal': fmt_num(vals[39]),
            'itr_total': fmt_num(vals[41]),
            'itr_closed': fmt_num(vals[42]),
            'itr_bal': fmt_num(vals[43]),
            'itr_pct': vals[44],
            'b1': fmt_num(vals[49]),  # CPP-1 blocking
            'b2': fmt_num(vals[50]),  # EIT blocking
            'b3': fmt_num(vals[51]),  # CPY/EACOP blocking
            're1': fmt_str(vals[52]),  # EIT remarks
            're2': fmt_str(vals[53]),  # CPP remarks
            're3': fmt_str(vals[55]),  # Walkdown/observation
        }
    
    wb_dpr.close()
    print(f"  RFC PROGRESS rows: {len(dpr_data)}")
else:
    print("  WARNING: No DPR SUMMARY file found")

print("\n" + "="*60)
print("STEP 4: Read cable data from PS5 Master tracker")
print("="*60)

cable_file = os.path.join(dpr_dir, 'PS5 Master tracker EIT Combined.xlsx')
cable_sheets = ['Electrical Cable Schedule', 'Instrument Cable Schedule', 'Telecom Cable Schedule', 'Trace Heating - Cable Schedule']
cable_labels = ['Electrical', 'Instrument', 'Telecom', 'Trace Heating']

CABLES = []  # [sid, disc, cable_tag, from_tag, from_desc, from_loc, to_tag, to_desc, to_loc, pulling_date, rfi_num, eit_date, rfi_date, punch]
sub_cable_count = defaultdict(int)

if os.path.exists(cable_file):
    wb_cab = openpyxl.load_workbook(cable_file, read_only=True, data_only=True)
    
    for si, (sheet_name, disc_label) in enumerate(zip(cable_sheets, cable_labels)):
        if sheet_name not in wb_cab.sheetnames:
            print(f"  WARNING: Sheet '{sheet_name}' not found")
            continue
        ws = wb_cab[sheet_name]
        
        # Find header row (row with 8+ non-empty cells)
        header_row = None
        for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
            non_empty = sum(1 for c in row if c is not None and str(c).strip())
            if non_empty >= 8:
                header_row = ri
                break
        if not header_row:
            print(f"  WARNING: No header found in {sheet_name}")
            continue
        
        # Read all rows after header
        for row in ws.iter_rows(min_row=header_row+1, values_only=True):
            vals = list(row)
            if not vals or not vals[1]:  # col 1 = Cable Tag
                continue
            cable_tag = str(vals[1] or '').strip()
            if not cable_tag or not cable_tag.startswith('PS5'):
                continue
            
            # Extract subsystem ID from cable tag: PS5-01-BE-0001-CL01 -> PS5-01
            parts = cable_tag.split('-')
            if len(parts) >= 2:
                sid = parts[0] + '-' + parts[1]
            else:
                sid = cable_tag
            
            # From side: col 10=Equipment Tag No., col 12=Description, col 14=Location
            from_tag = str(vals[10] or '').strip() if len(vals) > 10 else ''
            from_desc = str(vals[12] or '').strip() if len(vals) > 12 else ''
            from_loc = str(vals[14] or '').strip() if len(vals) > 14 else ''
            # To side: col 17=Equipment Tag No., col 19=Description, col 21=Location
            to_tag = str(vals[17] or '').strip() if len(vals) > 17 else ''
            to_desc = str(vals[19] or '').strip() if len(vals) > 19 else ''
            to_loc = str(vals[21] or '').strip() if len(vals) > 21 else ''
            
            # Try to match from/to tags against ITR asset tag map
            tag_sid = tag_to_sub.get(from_tag) or tag_to_sub.get(to_tag)
            if tag_sid:
                sid = tag_sid
            
            # Dates
            pulling_date = ''
            rfi_num = ''
            eit_date = ''
            rfi_date = ''
            punch = ''
            
            if disc_label == 'Trace Heating':
                # Trace Heating has different col layout
                pulling_date = fmt_date(vals[30]) if len(vals) > 30 else ''
                rfi_num = str(vals[30] or '').strip() if len(vals) > 30 else ''
                rfi_date = fmt_date(vals[31]) if len(vals) > 31 else ''
                punch = str(vals[39] or '').strip() if len(vals) > 39 else ''
            else:
                pulling_date = fmt_date(vals[30]) if len(vals) > 30 else ''
                rfi_num = str(vals[31] or '').strip() if len(vals) > 31 else ''
                eit_date = fmt_date(vals[38]) if len(vals) > 38 else ''
                rfi_date = fmt_date(vals[39]) if len(vals) > 39 else ''
            
            CABLES.append([
                sid,           # 0: subsystem id
                disc_label,    # 1: discipline
                cable_tag,     # 2: cable tag
                from_tag,      # 3: from equipment tag
                from_desc,     # 4: from description
                from_loc,      # 5: from location
                to_tag,        # 6: to equipment tag
                to_desc,       # 7: to description
                to_loc,        # 8: to location
                pulling_date,  # 9: pulling date
                rfi_num,       # 10: rfi num
                eit_date,      # 11: eit date
                rfi_date,      # 12: rfi date
                punch,         # 13: punch description
            ])
            sub_cable_count[sid] += 1
    
    wb_cab.close()
    print(f"  CABLES: {len(CABLES)} cables")
    print(f"  Subsystems with cables: {len(sub_cable_count)}")
else:
    print(f"  WARNING: Cable file not found: {cable_file}")

print("\n" + "="*60)
print("STEP 5: Build SUBS from unique subsystems in ITR")
print("="*60)

all_subs = set()
for r in itr_raw:
    if r[9]: all_subs.add(r[9])
for r in punch_raw:
    if r[9]: all_subs.add(r[9])

SUBS = []
for sid in sorted(all_subs):
    name = sub_name_map.get(sid, sid)
    if not name or name == sid:
        name = sid
    SUBS.append({"name": name, "sid": sid})

print(f"  SUBS: {len(SUBS)} subsystems")

print("\n" + "="*60)
print("STEP 6: Build RFC from ITR + DPR SUMMARY")
print("="*60)

RFK = ['B', 'E', 'H', 'I', 'M', 'P', 'S', 'T']
disc_name = {'B': 'Building', 'E': 'Electrical', 'H': 'HVAC', 'I': 'Instrumentation',
             'M': 'Mechanical', 'P': 'Piping & Vessel', 'S': 'Safety', 'T': 'Telecom'}

# Aggregate ITR per subsystem per discipline
sub_disc = defaultdict(lambda: defaultdict(lambda: {'total': 0, 'closed': 0}))
sub_itr_total = defaultdict(lambda: {'total': 0, 'closed': 0})

for r in itr_raw:
    sid = r[9]
    disc = r[2]
    is_done = r[5].strip().lower().startswith('comp') or r[5].strip().lower() == 'closed'
    if sid and disc:
        sub_disc[sid][disc]['total'] += 1
        if is_done: sub_disc[sid][disc]['closed'] += 1
    if sid:
        sub_itr_total[sid]['total'] += 1
        if is_done: sub_itr_total[sid]['closed'] += 1

# Preserve existing OVR overrides for RFC fields
rfc_ovr = OVR.get('RFC PROGRESS', {})

RFC = []
for s in SUBS:
    sid = s['sid']
    d = []
    for k in RFK:
        t = int(sub_disc[sid][k]['total'])
        c = int(sub_disc[sid][k]['closed'])
        d.append({'t': t, 'c': c})
    
    it = sub_itr_total[sid]['total']
    ic = sub_itr_total[sid]['closed']
    ib = it - ic
    
    sT = sum(x['t'] for x in d)
    sC = sum(x['c'] for x in d)
    itrp = round(ic / it * 100) if it > 0 else ''
    
    # Use DPR data if available, otherwise keep overrides
    dd = dpr_data.get(sid, {})
    ro = rfc_ovr.get(sid, {})
    
    # TOTAL % from DPR col I (0..1 fraction); fallback to ITR aggregation
    dpr_tot = dd.get('tot_pct')
    if dpr_tot not in (None, '') and isinstance(dpr_tot, (int, float)):
        tot = round(float(dpr_tot) * 100)
    else:
        tot = round(sC / sT * 100) if sT > 0 else ''
    
    RFC.append({
        'sid': sid,
        'name': s['name'],
        'prio': dd.get('priority') or ro.get('Priority', ''),
        'bhm': dd.get('bhm') or ro.get('RFC BHMPS', ''),
        'eit': dd.get('eit') or ro.get('RFC EIT', ''),
        'base': dd.get('base') or ro.get('RFC BASE', ''),
        'rec': dd.get('rec') or ro.get('Recovery', ''),
        'signed': dd.get('signed') or ro.get('SIGNED', ''),
        'mile': dd.get('mile') or ro.get('Milestone', ''),
        'tot': tot,
        'itrp': itrp,
        'b1': dd.get('b1') or int(ro.get('CPP-1', 0) or 0),
        'b2': dd.get('b2') or int(ro.get('EIT', 0) or 0),
        'b3': dd.get('b3') or int(ro.get('EACOP', 0) or 0),
        're1': dd.get('re1') or ro.get('REMARK EACOP', ''),
        're2': dd.get('re2') or ro.get('REMARK CPP-EIT', ''),
        're3': dd.get('re3') or ro.get('REMARK CPP-1', ''),
        'wd': dd.get('re3') or ro.get('WALKDOWN STATUS', ''),
        'mile2': '',
        'd': d,
        'it': it,
        'ic': ic,
        'ib': ib
    })

print(f"  RFC: {len(RFC)} subsystems")
print(f"  Total ITRs: {sum(r['it'] for r in RFC)}, Closed: {sum(r['ic'] for r in RFC)}")

# Show sample with DPR data
sample_count = 0
for r in RFC:
    if r['prio'] and sample_count < 3:
        print(f"  Sample: {r['sid']} prio={r['prio']} bhm={r['bhm']} mile={r['mile']} b1={r['b1']} b2={r['b2']} b3={r['b3']}")
        sample_count += 1

print("\n" + "="*60)
print("STEP 7: Build ITRT")
print("="*60)

def get_group(sid):
    m = re.match(r'PS5-(\d{2})-', sid)
    if not m: return 'Other'
    n = int(m.group(1))
    if 1 <= n <= 10: return 'CPP'
    if 11 <= n <= 29: return 'CPP'
    if 30 <= n <= 39: return 'CPP'
    if 40 <= n <= 49: return 'CPP'
    if 50 <= n <= 59: return 'CPP'
    if 60 <= n <= 69: return 'EIT'
    if 70 <= n <= 79: return 'EIT'
    if 80 <= n <= 89: return 'EACOP'
    if 90 <= n <= 99: return 'EACOP'
    return 'Other'

group_disc = defaultdict(lambda: defaultdict(lambda: {'total': 0, 'closed': 0}))

for r in itr_raw:
    sid = r[9]
    disc = r[2]
    if not sid or not disc: continue
    group = get_group(sid)
    is_done = r[5].strip().lower().startswith('comp') or r[5].strip().lower() == 'closed'
    group_disc[group][disc]['total'] += 1
    if is_done:
        group_disc[group][disc]['closed'] += 1

ITRT = []
group_order = ['CPP', 'EIT', 'EACOP']
disc_order = ['B', 'E', 'H', 'I', 'M', 'P', 'S', 'T']

for grp in group_order:
    grp_total = 0
    grp_closed = 0
    for d in disc_order:
        dn = disc_name.get(d, d)
        t = group_disc[grp][d]['total']
        c = group_disc[grp][d]['closed']
        ratio = c / t if t > 0 else 0
        ITRT.append([grp, dn, t, c, ratio, t - c])
        grp_total += t
        grp_closed += c
    ratio = grp_closed / grp_total if grp_total > 0 else 0
    ITRT.append([grp, 'Total', grp_total, grp_closed, ratio, grp_total - grp_closed])

gt = sum(r['it'] for r in RFC)
gc = sum(r['ic'] for r in RFC)
ITRT.append(['Grand', 'Total', gt, gc, gc/gt if gt > 0 else 0, gt - gc])

print(f"  ITRT rows: {len(ITRT)}")

print("\n" + "="*60)
print("STEP 8: Build PUNT")
print("="*60)

cat_order = ['A', 'B', 'C']
disc_order_p = ['B', 'E', 'H', 'I', 'M', 'P', 'S', 'T']
disc_name_p = {'B': 'Building', 'E': 'Electrical', 'H': 'HVAC', 'I': 'Instrumentation',
               'M': 'Mechanical', 'P': 'Piping & Vessel', 'S': 'Safety', 'T': 'Telecom',
               'Miscellaneous': 'Miscellaneous'}

PUNT = []
for d in disc_order_p:
    dn = disc_name_p.get(d, d)
    row = [dn]
    for cat in cat_order:
        t = punch_disc_cat[d][cat]['total']
        c = punch_disc_cat[d][cat]['closed']
        row.extend([t, c, t - c])
    PUNT.append(row)

grand = ['Grand Total']
for cat in cat_order:
    t = sum(punch_disc_cat[d][cat]['total'] for d in punch_disc_cat)
    c = sum(punch_disc_cat[d][cat]['closed'] for d in punch_disc_cat)
    grand.extend([t, c, t - c])
PUNT.append(grand)

print(f"  PUNT rows: {len(PUNT)}")

print("\n" + "="*60)
print("STEP 9: Build new data line and replace in HTML")
print("="*60)

def fmt_arr(arr):
    return json.dumps(arr, ensure_ascii=False, separators=(',', ':'))

subs_str = fmt_arr(SUBS)
punch_str = fmt_arr(punch_raw)
itr_str = fmt_arr(itr_raw)
rfc_str = fmt_arr(RFC)
itrt_str = fmt_arr(ITRT)
punt_str = fmt_arr(PUNT)
miles_str = fmt_arr(MILES)
cables_str = fmt_arr(CABLES)

new_line227 = f'const SUBS={subs_str},PUNCH={punch_str},ITR={itr_str},RFC={rfc_str},ITRT={itrt_str},PUNT={punt_str},MILES={miles_str},CABLES={cables_str};'

print(f"\n  VERIFICATION:")
print(f"  ITR: {len(itr_raw)} tasks")
print(f"  PUNCH: {len(punch_raw)} items")
print(f"  SUBS: {len(SUBS)} subsystems")
print(f"  RFC: {len(RFC)} subsystems")
print(f"  ITRT rows: {len(ITRT)}")
print(f"  PUNT rows: {len(PUNT)}")
print(f"  MILES: {len(MILES)}")
print(f"  CABLES: {len(CABLES)} cables")
print(f"  DPR data: {len(dpr_data)} subsystems with workflow fields")

# Replace data line in HTML
old_data_size = len(data_line)
print(f"\n  Old data line size: {old_data_size} chars")
print(f"  New data line size: {len(new_line227)} chars")

lines[data_line_idx] = new_line227

# STEP 10: Auto-detect report date
print("\n" + "="*60)
print("STEP 10: Auto-detect report date")
print("="*60)

if dpr_files:
    dpr_name = os.path.basename(dpr_files[0])
    date_match = re.search(r'(\d{2}-\d{2}-\d{2})', dpr_name)
    if date_match:
        report_date = date_match.group(1)
        source_name = dpr_name.replace('.xlsx', '')
        print(f"  Date: {report_date}")
        
        for i, ln in enumerate(lines):
            if 'Report Date:' in ln and 'Source:' in ln:
                new_meta = f'<div class="meta">Report Date: {report_date}    |    Source: {source_name}.xlsx</div>'
                lines[i] = ln.replace(
                    re.search(r'<div class="meta">.*?</div>', ln).group(0),
                    new_meta
                )
                print(f"  Updated header at line {i+1}")
                break

new_html = '\n'.join(lines)

out_path = 'C:/Users/mylap/OneDrive/Desktop/PS5-COMPLETION-PLATFORM/index.html'
with open(out_path, 'w', encoding='utf-8') as f:
    f.write(new_html)

print(f"\n  Written to: {out_path}")
print(f"  New file size: {os.path.getsize(out_path)} bytes")

print("\n" + "="*60)
print("DONE! Data rebuilt from XLSX sources.")
print("="*60)
